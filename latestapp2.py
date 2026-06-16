from __future__ import annotations

# ─────────────────────────────────────────────────────────────────────────────
#  Cash Flow Manager — CFO Edition
#  Improvements over original:
#    P1: Configurable collection probabilities · Min-cash alert banner ·
#        Multi-region holiday calendar (Karnataka 2026-27, Maharashtra,
#        Delhi NCR) · TDS deduction on collections · Typo fixes
#    P2: OD / credit-line KPIs (sanctioned limit, headroom, interest est.) ·
#        Audit trail (load timestamp + file context)
#    P3: CFO narrative commentary boxes · Variance commentary in Actual vs
#        Budget · Commentary included in Excel export
# ─────────────────────────────────────────────────────────────────────────────

import ast
import hashlib
import html
import json
from calendar import monthrange
from datetime import date, datetime, timedelta
from io import BytesIO
from numbers import Number
from pathlib import Path
import re
from typing import Any
from urllib.parse import urlencode

import numpy as np
import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter


APP_TITLE = "Cash Flow Manager"
DEFAULT_SOURCE = Path(__file__).with_name("Data Source.xlsx")
TODAY = pd.Timestamp(date.today()).normalize()
DATE_DISPLAY_FORMAT = "%d-%m-%Y"

# ── Default collection probabilities (overridden via sidebar sliders) ─────────
DEFAULT_BASE_COLLECTION_PROBABILITIES: dict[str, float] = {
    "0-30": 1.00,
    "30-60": 1.00,
    "60+": 1.00,
}
DEFAULT_WORST_COLLECTION_PROBABILITIES: dict[str, float] = {
    "0-30": 0.80,
    "30-60": 0.80,
    "60+": 0.80,
}
COLLECTION_DATE_OFFSETS: dict[str, int] = {
    "0-30": 10,
    "30-60": 35,
    "60+": 70,
}

# ── Default model assumption constants ────────────────────────────────────────
DEFAULT_MIN_CASH_THRESHOLD: float = 500_000.0   # ₹5 lakh minimum cash floor
DEFAULT_TDS_RATE: float = 0.0                    # TDS % on gross collections
DEFAULT_OD_LIMIT: float = 0.0                    # Sanctioned OD / CC limit
DEFAULT_OD_RATE_PA: float = 12.0                 # OD interest rate % p.a.
DEFAULT_OPENING_OD_UTILIZATION: float = 0.0      # OD outstanding brought forward

# ── Holiday calendar ──────────────────────────────────────────────────────────
STATE_BANK_HOLIDAYS: dict[str, dict[int, dict[str, str]]] = {
    "Karnataka (Bengaluru)": {
        2026: {
            "2026-01-14": "Makara Sankranti / Uttarayana Punyakala",
            "2026-01-26": "Republic Day",
            "2026-03-20": "Ugadi",
            "2026-03-30": "Mahavir Jayanti (Karnataka observed date)",
            "2026-04-01": "Annual Accounts Closing (Bank Holiday)",
            "2026-04-03": "Good Friday",
            "2026-04-14": "Dr. B. R. Ambedkar Jayanti",
            "2026-04-20": "Basava Jayanthi / Akshaya Tritiya",
            "2026-05-01": "May Day / Buddha Purnima",
            "2026-05-27": "Bakrid / Eid al-Adha - tentative",
            "2026-06-26": "Muharram - tentative",
            "2026-08-15": "Independence Day",
            "2026-08-25": "Id-E-Milad - tentative",
            "2026-09-14": "Ganesh Chaturthi",
            "2026-10-02": "Mahatma Gandhi Jayanti",
            "2026-10-20": "Maha Navami / Ayudha Puja / Vijayadashami",
            "2026-10-26": "Maharshi Valmiki Jayanti",
            "2026-11-09": "Deepavali Holiday",
            "2026-11-27": "Kanakadasa Jayanthi",
            "2026-12-25": "Christmas Day",
        },
        2027: {
            "2027-01-14": "Makara Sankranti",
            "2027-01-26": "Republic Day",
            "2027-04-01": "Annual Accounts Closing (Bank Holiday)",
            "2027-04-14": "Dr. B. R. Ambedkar Jayanti",
            "2027-04-23": "Good Friday - tentative",
            "2027-05-01": "May Day / Basava Jayanthi",
            "2027-08-15": "Independence Day",
            "2027-09-02": "Ganesh Chaturthi - tentative",
            "2027-10-02": "Mahatma Gandhi Jayanti",
            "2027-10-20": "Vijaya Dashami - tentative",
            "2027-11-05": "Deepavali - tentative",
            "2027-12-25": "Christmas Day",
        },
    },
    "Maharashtra (Mumbai)": {
        2026: {
            "2026-01-26": "Republic Day",
            "2026-02-19": "Chhatrapati Shivaji Maharaj Jayanti",
            "2026-03-17": "Holi (Second Day)",
            "2026-03-20": "Gudi Padwa",
            "2026-04-01": "Annual Accounts Closing (Bank Holiday)",
            "2026-04-03": "Good Friday",
            "2026-04-14": "Dr. B. R. Ambedkar Jayanti",
            "2026-05-01": "Maharashtra Day / Buddha Purnima",
            "2026-05-27": "Eid-Ul-Adha - tentative",
            "2026-06-26": "Muharram - tentative",
            "2026-08-15": "Independence Day",
            "2026-08-25": "Id-E-Milad - tentative",
            "2026-09-14": "Ganesh Chaturthi",
            "2026-10-02": "Mahatma Gandhi Jayanti",
            "2026-10-20": "Dussehra",
            "2026-11-09": "Deepavali Holiday",
            "2026-12-25": "Christmas Day",
        },
        2027: {
            "2027-01-26": "Republic Day",
            "2027-02-19": "Chhatrapati Shivaji Maharaj Jayanti",
            "2027-04-01": "Annual Accounts Closing (Bank Holiday)",
            "2027-04-14": "Dr. B. R. Ambedkar Jayanti",
            "2027-05-01": "Maharashtra Day",
            "2027-08-15": "Independence Day",
            "2027-10-02": "Mahatma Gandhi Jayanti",
            "2027-12-25": "Christmas Day",
        },
    },
    "Delhi NCR": {
        2026: {
            "2026-01-26": "Republic Day",
            "2026-02-26": "Maha Shivaratri - tentative",
            "2026-03-17": "Holi",
            "2026-04-01": "Annual Accounts Closing (Bank Holiday)",
            "2026-04-03": "Good Friday",
            "2026-04-14": "Dr. B. R. Ambedkar Jayanti",
            "2026-04-20": "Ram Navami - tentative",
            "2026-05-01": "Buddha Purnima",
            "2026-05-27": "Eid-Ul-Adha - tentative",
            "2026-06-26": "Muharram - tentative",
            "2026-08-15": "Independence Day",
            "2026-08-25": "Id-E-Milad - tentative",
            "2026-10-02": "Mahatma Gandhi Jayanti",
            "2026-10-20": "Dussehra",
            "2026-10-27": "Diwali - tentative",
            "2026-11-05": "Guru Nanak Jayanti - tentative",
            "2026-12-25": "Christmas Day",
        },
        2027: {
            "2027-01-26": "Republic Day",
            "2027-04-01": "Annual Accounts Closing (Bank Holiday)",
            "2027-08-15": "Independence Day",
            "2027-10-02": "Mahatma Gandhi Jayanti",
            "2027-12-25": "Christmas Day",
        },
    },
}

TOTAL_ROWS: set[str] = {
    "Inflows",
    "Total Inflows",
    "Cash Outflows",
    "Total Outflows",
    "Net Movement of Cash",
    "Opening Balance incl. OD Opening O/S",
    "Closing Balance",
}

PREFERRED_OUTFLOW_ORDER: list[str] = [
    "Employee Salaries",
    "Mcom Salaries",
    "Statutory Dues",
    "Fixed Expenses",
    "Vendor Payment",
    "Reimbursements",   # Fixed: was "Reimbursments"
]
DEFAULT_RECEIVABLE_EXPORT_COLUMNS: list[str] = [
    "Date",
    "Ref. No.",
    "Clients",
    "Pending Amt",
    "Ageing",
    "Tentative Collection Date",
    "Due date",
    "Billing Type",
]
DEFAULT_OUTFLOW_EXPORT_COLUMNS: list[str] = [
    "Payment Date",
    "Description",
    "Amount",
    "Party Name",
]
INFLOW_SHEET_NAME_HINTS: tuple[str, ...] = (
    "receivable",
    "inflow",
    "receipt",
    "receipts",
    "collection",
    "collections",
    "customeradvance",
    "advancefromcustomer",
    "advancereceived",
    "clientreceipt",
)
OUTFLOW_SHEET_NAME_HINTS: tuple[str, ...] = (
    "payment",
    "payments",
    "payable",
    "expense",
    "expenses",
    "vendor",
    "salary",
    "salaries",
    "dues",
    "reimbursement",
    "reimbursements",
    "advancevendor",
    "vendoradvance",
)

CELL_REF_PATTERN = re.compile(r"\$?([A-Z]{1,3})\$?(\d+)")
WEEKLY_DRILLDOWN_QUERY_KEYS = (
    "weekly_drill_scope",
    "weekly_drill_week",
    "weekly_drill_line",
)
VIEW_QUERY_KEY = "cashflow_view"
FORECAST_MONTH_QUERY_KEY = "forecast_month"
BALANCE_DATE_QUERY_KEY = "balance_as_of"
DEFAULT_NAV_VIEW = "Executive Summary"
GROUP_CASHFLOW_VIEW = "Group Cashflow"
UPLOAD_CACHE_DIR = Path(__file__).with_name(".cashflow_upload_cache")

ALLOWED_AST_BINARY_OPERATORS: tuple[type[ast.operator], ...] = (
    ast.Add, ast.Sub, ast.Mult, ast.Div, ast.Mod, ast.Pow,
)
ALLOWED_AST_UNARY_OPERATORS: tuple[type[ast.unaryop], ...] = (
    ast.UAdd, ast.USub,
)


# ══════════════════════════════════════════════════════════════════════════════
#  UTILITY HELPERS
# ══════════════════════════════════════════════════════════════════════════════

def standardize_label(value: str) -> str:
    return "".join(ch.lower() for ch in str(value).strip() if ch.isalnum())


def order_outflow_sheets(sheet_names: list[str]) -> list[str]:
    preferred_rank = {standardize_label(name): idx for idx, name in enumerate(PREFERRED_OUTFLOW_ORDER)}
    return sorted(
        sheet_names,
        key=lambda name: (preferred_rank.get(standardize_label(name), len(preferred_rank)), standardize_label(name)),
    )


def order_sheet_names(sheet_names: list[str]) -> list[str]:
    seen: set[str] = set()
    ordered: list[str] = []
    for sheet_name in sheet_names:
        key = standardize_label(sheet_name)
        if not key or key in seen:
            continue
        seen.add(key)
        ordered.append(str(sheet_name))
    return ordered


def find_column(columns: list[str], options: list[str]) -> str | None:
    normalized = {col: standardize_label(col) for col in columns}
    targets = [standardize_label(option) for option in options if standardize_label(option)]
    for target in targets:
        for col, label in normalized.items():
            if label == target:
                return col
    for target in targets:
        for col, label in normalized.items():
            if target and target in label:
                return col
    return None


def to_numeric(series: pd.Series) -> pd.Series:
    return pd.to_numeric(series, errors="coerce").fillna(0.0)


def to_datetime(series: pd.Series) -> pd.Series:
    return pd.to_datetime(series, errors="coerce").dt.normalize()


def build_summary_label(
    df: pd.DataFrame,
    preferred_column: str | None,
    fallback_column: str | None,
    default_value: str,
) -> pd.Series:
    label = pd.Series("", index=df.index, dtype="object")
    for column in [preferred_column, fallback_column]:
        if column and column in df.columns:
            values = df[column].fillna("").astype(str).str.strip()
            label = label.mask(label.eq(""), values)
    return label.mask(label.eq(""), default_value)


def detect_receivables_columns(columns: list[str]) -> dict[str, str | None]:
    return {
        "date": find_column(columns, ["invoice date", "collection date", "receipt date", "actual date", "date"]),
        "amount": find_column(columns, ["net receipt", "receipt amount", "pending amt", "gross amt", "basic amt", "amount"]),
        "client": find_column(columns, ["clients", "customer", "client", "party name", "party", "counterparty", "particulars", "particualrs"]),
        "reference": find_column(columns, ["ref no", "invoice no", "invoice number", "inv number"]),
        "ageing": find_column(columns, ["ageing", "aging"]),
        "tentative_date": find_column(columns, ["tentative collection date", "tentative date", "expected collection date"]),
        "due_date": find_column(columns, ["due date"]),
        "billing_type": find_column(columns, ["billing type"]),
    }


def detect_outflow_columns(columns: list[str]) -> dict[str, str | None]:
    return {
        "planned_date": find_column(columns, ["payment date", "plan payment date", "date", "invoice date"]),
        "actual_date": find_column(columns, ["actual date"]),
        "amount": find_column(columns, ["payable", "amount", "balance", "gross", "net amount", "net payable"]),
        "vendor": find_column(columns, ["vendors name as per tally", "vendor name", "vendor", "supplier", "payee", "party name", "party"]),
        "description": find_column(
            columns,
            ["description", "expense", "nature of payment", "particulars", "particualrs", "invoice no", "inv. number", "inv number"],
        ),
    }


def classify_sheet_schema(sheet_name: str, df: pd.DataFrame) -> dict[str, Any]:
    columns = list(df.columns)
    receivables_map = detect_receivables_columns(columns)
    outflow_map = detect_outflow_columns(columns)
    normalized_sheet_name = standardize_label(sheet_name)
    receivables_signals = sum(
        1 for key in ["client", "reference", "ageing", "tentative_date", "due_date", "billing_type"]
        if receivables_map.get(key)
    )
    outflow_signals = sum(1 for key in ["vendor", "description", "actual_date"] if outflow_map.get(key))
    has_receivables_name = any(hint in normalized_sheet_name for hint in INFLOW_SHEET_NAME_HINTS)
    has_outflow_name = any(hint in normalized_sheet_name for hint in OUTFLOW_SHEET_NAME_HINTS)
    receivables_date_col = receivables_map.get("tentative_date") or receivables_map.get("due_date") or receivables_map.get("date")
    outflow_date_col = outflow_map.get("planned_date") or outflow_map.get("actual_date")
    inflow_amount_present = bool(receivables_map.get("amount"))
    outflow_amount_present = bool(outflow_map.get("amount"))
    looks_like_receivables = inflow_amount_present and bool(receivables_date_col)
    looks_like_outflow = outflow_amount_present and bool(outflow_date_col)
    strong_receivables = has_receivables_name or (receivables_signals >= 2 and not has_outflow_name)
    strong_outflow = has_outflow_name or bool(outflow_map.get("vendor") or outflow_map.get("description")) or (
        outflow_signals >= 1 and not has_receivables_name
    )

    if looks_like_receivables and strong_receivables and not strong_outflow:
        return {"role": "receivables", "column_map": receivables_map, "columns": columns}
    if looks_like_outflow and strong_outflow and not strong_receivables:
        return {"role": "outflow", "column_map": outflow_map, "columns": columns}

    # For user-added operational tabs, default any date+amount sheet to outflow
    # unless it has strong receivable-style signals.
    if looks_like_outflow and not strong_receivables:
        return {"role": "outflow", "column_map": outflow_map, "columns": columns}
    if looks_like_receivables:
        return {"role": "receivables", "column_map": receivables_map, "columns": columns}

    return {"role": "ignored", "column_map": {}, "columns": columns}


def reset_source_pointer(source: Any) -> None:
    if hasattr(source, "seek"):
        source.seek(0)


def build_month_start(value: Any) -> pd.Timestamp:
    return pd.Timestamp(value).to_period("M").to_timestamp()


def build_month_end(value: Any) -> pd.Timestamp:
    return build_month_start(value) + pd.offsets.MonthEnd(1)


def format_date(value: Any) -> str:
    if value is None or pd.isna(value):
        return ""
    return pd.Timestamp(value).strftime(DATE_DISPLAY_FORMAT)


def month_label(value: Any) -> str:
    return pd.Timestamp(value).strftime("%b %Y")


def default_balance_as_of_date(selected_month: pd.Timestamp) -> pd.Timestamp:
    month_start = build_month_start(selected_month)
    month_end = build_month_end(selected_month)
    if month_start <= TODAY <= month_end:
        return TODAY
    return month_start


def short_date_label(value: pd.Timestamp) -> str:
    return format_date(value)


def full_date_with_day(value: pd.Timestamp) -> str:
    return pd.Timestamp(value).strftime("%a %d %b %Y")


def compact_day_date_label(value: pd.Timestamp) -> str:
    return pd.Timestamp(value).strftime("%a %d %b")


def format_date_columns(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df.copy()
    display_df = df.copy()
    for column in display_df.columns:
        series = display_df[column]
        if pd.api.types.is_datetime64_any_dtype(series):
            display_df[column] = series.map(format_date)
            continue
        if series.dtype == "object":
            non_null = series.dropna()
            if not non_null.empty and non_null.map(
                lambda value: isinstance(value, (pd.Timestamp, datetime, date))
            ).all():
                display_df[column] = series.map(format_date)
    return display_df


def get_query_params() -> dict[str, str]:
    query_params_api = getattr(st, "query_params", None)
    if query_params_api is not None:
        try:
            return {
                key: values[-1] if isinstance(values, list) else str(values)
                for key, values in query_params_api.items()
            }
        except Exception:
            pass
    getter = getattr(st, "experimental_get_query_params", None)
    if callable(getter):
        raw_params = getter()
        return {
            key: values[-1] if isinstance(values, list) else str(values)
            for key, values in raw_params.items()
        }
    return {}


def set_query_params(params: dict[str, str]) -> None:
    query_params_api = getattr(st, "query_params", None)
    if query_params_api is not None:
        try:
            query_params_api.clear()
            for key, value in params.items():
                query_params_api[key] = value
            return
        except Exception:
            pass
    setter = getattr(st, "experimental_set_query_params", None)
    if callable(setter):
        setter(**params)


def clear_weekly_drilldown_query_params() -> None:
    params = get_query_params()
    for key in WEEKLY_DRILLDOWN_QUERY_KEYS:
        params.pop(key, None)
    set_query_params(params)


def sync_forecast_month_query_param(selected_month: pd.Timestamp) -> None:
    params = get_query_params()
    selected_value = build_month_start(selected_month).strftime("%Y-%m")
    if params.get(FORECAST_MONTH_QUERY_KEY) == selected_value:
        return
    params[FORECAST_MONTH_QUERY_KEY] = selected_value
    set_query_params(params)


def build_weekly_drilldown_href(
    scope: str,
    week_key: str,
    line_item: str,
    selected_month: pd.Timestamp,
    balance_as_of_date: pd.Timestamp,
) -> str:
    params = get_query_params()
    for key in WEEKLY_DRILLDOWN_QUERY_KEYS:
        params.pop(key, None)
    params.update({
        "weekly_drill_scope": scope,
        "weekly_drill_week": week_key,
        "weekly_drill_line": line_item,
        VIEW_QUERY_KEY: GROUP_CASHFLOW_VIEW,
        FORECAST_MONTH_QUERY_KEY: build_month_start(selected_month).strftime("%Y-%m"),
        BALANCE_DATE_QUERY_KEY: pd.Timestamp(balance_as_of_date).strftime("%Y-%m-%d"),
    })
    return f"?{urlencode(params)}#weekly-drilldown-target"


def has_weekly_drilldown_query() -> bool:
    params = get_query_params()
    return all(params.get(key) for key in WEEKLY_DRILLDOWN_QUERY_KEYS)


def get_requested_view(default_view: str = DEFAULT_NAV_VIEW) -> str:
    params = get_query_params()
    requested_view = params.get(VIEW_QUERY_KEY, "").strip()
    return requested_view or default_view


def upload_cache_path(kind: str, suffix: str) -> Path:
    UPLOAD_CACHE_DIR.mkdir(exist_ok=True)
    return UPLOAD_CACHE_DIR / f"{kind}{suffix}"


def build_named_bytes_io(payload: bytes, name: str) -> BytesIO:
    buffer = BytesIO(payload)
    buffer.name = name
    return buffer


def persist_uploaded_file(kind: str, uploaded: Any, load_time: str | None = None) -> None:
    binary_path = upload_cache_path(kind, ".bin")
    meta_path = upload_cache_path(kind, ".json")
    if uploaded is None:
        if binary_path.exists():
            binary_path.unlink()
        if meta_path.exists():
            meta_path.unlink()
        return

    reset_source_pointer(uploaded)
    payload = uploaded.read()
    reset_source_pointer(uploaded)
    binary_path.write_bytes(payload)
    meta_path.write_text(
        json.dumps({
            "name": getattr(uploaded, "name", f"{kind}.xlsx"),
            "load_time": load_time,
        }),
        encoding="utf-8",
    )


def load_cached_upload(kind: str) -> tuple[BytesIO | None, dict[str, Any]]:
    binary_path = upload_cache_path(kind, ".bin")
    meta_path = upload_cache_path(kind, ".json")
    if not binary_path.exists():
        return None, {}
    meta: dict[str, Any] = {}
    if meta_path.exists():
        try:
            meta = json.loads(meta_path.read_text(encoding="utf-8"))
        except Exception:
            meta = {}
    buffer = build_named_bytes_io(
        binary_path.read_bytes(),
        meta.get("name", f"{kind}.xlsx"),
    )
    return buffer, meta


def normalize_billing_bucket(value: Any) -> str:
    label = str(value or "").strip().lower()
    if "adv" in label:
        return "Advance"
    return "Receivables"


def classify_weekly_drilldown_line_item(
    line_item: str,
    outflow_categories: list[str],
) -> dict[str, str] | None:
    if line_item == "Total Inflows":
        return {"section": "inflows", "mode": "all", "label": "Collections"}
    if line_item == "   Advance":
        return {"section": "inflows", "mode": "advance", "label": "Advance Collections"}
    if line_item == "   Receivables":
        return {"section": "inflows", "mode": "receivables", "label": "Receivables"}
    if line_item == "      Due Within Week":
        return {"section": "inflows", "mode": "due_within_week", "label": "Due Within Week"}
    if line_item == "      Overdue":
        return {"section": "inflows", "mode": "overdue", "label": "Overdue"}
    if line_item == "Total Outflows":
        return {"section": "outflows", "mode": "all", "label": "Total Outflows"}
    category_name = line_item.strip()
    normalized_categories = {
        standardize_label(category): str(category)
        for category in outflow_categories
        if str(category).strip()
    }
    matched_category = normalized_categories.get(standardize_label(category_name))
    if matched_category:
        return {
            "section": "outflows",
            "mode": "category",
            "category": matched_category,
            "label": matched_category,
        }
    return None


def file_md5(source: Any) -> str:
    """Return a short MD5 hex digest of the uploaded file bytes."""
    try:
        reset_source_pointer(source)
        digest = hashlib.md5(source.read()).hexdigest()[:8].upper()
        reset_source_pointer(source)
        return digest
    except Exception:
        return "N/A"


EDIT_HISTORY_CONFIG: dict[str, dict[str, Any]] = {
    "receivables": {
        "label": "Inflows",
        "line_label": "receivable line",
        "key_columns": ["counterparty", "invoice_date", "due_date"],
        "detail_columns": ["amount", "tentative_collection_date", "ageing", "aging_bucket", "billing_bucket"],
    },
    "outflows": {
        "label": "Outflows",
        "line_label": "outflow entry",
        "key_columns": ["sheet_name", "vendor_name", "date"],
        "detail_columns": ["amount", "description"],
    },
}


def _edit_state_key(dataset_key: str) -> str:
    return f"edited_{dataset_key}"


def _edit_history_key(dataset_key: str) -> str:
    return f"_edit_history_{dataset_key}"


def _edit_history_index_key(dataset_key: str) -> str:
    return f"_edit_history_{dataset_key}_index"


def _edit_history_log_key(dataset_key: str) -> str:
    return f"_edit_history_{dataset_key}_log"


def _normalize_history_scalar(value: Any) -> Any:
    if value is None or pd.isna(value):
        return ""
    if isinstance(value, (pd.Timestamp, datetime, date)):
        return pd.Timestamp(value).strftime("%Y-%m-%d")
    if isinstance(value, Number):
        return round(float(value), 2)
    return str(value).strip()


def _normalize_history_df(df: pd.DataFrame) -> pd.DataFrame:
    normalized = df.copy()
    for column in normalized.columns:
        if pd.api.types.is_datetime64_any_dtype(normalized[column]):
            normalized[column] = (
                pd.to_datetime(normalized[column], errors="coerce")
                .dt.strftime("%Y-%m-%d")
                .fillna("")
            )
        else:
            normalized[column] = normalized[column].map(_normalize_history_scalar)
    return normalized.fillna("")


def _build_history_index(df: pd.DataFrame, dataset_key: str) -> pd.DataFrame:
    normalized = _normalize_history_df(df)
    if normalized.empty:
        normalized["_history_key"] = pd.Series(dtype="object")
        return normalized.set_index("_history_key")
    config = EDIT_HISTORY_CONFIG[dataset_key]
    key_columns = [column for column in config["key_columns"] if column in normalized.columns]
    if not key_columns:
        key_columns = normalized.columns.tolist()
    base_key = normalized[key_columns].astype(str).agg(" | ".join, axis=1).replace("", "<blank>")
    sequence = base_key.groupby(base_key, sort=False).cumcount().astype(str)
    normalized["_history_key"] = base_key + " #" + sequence
    return normalized.set_index("_history_key", drop=True)


def edit_frames_equal(left_df: pd.DataFrame, right_df: pd.DataFrame) -> bool:
    all_columns = sorted(set(left_df.columns).union(right_df.columns))
    left_normalized = _normalize_history_df(left_df.reindex(columns=all_columns))
    right_normalized = _normalize_history_df(right_df.reindex(columns=all_columns))
    return left_normalized.equals(right_normalized)


def _format_history_field_name(field_name: str) -> str:
    return field_name.replace("_", " ").title()


def _format_history_value(field_name: str, value: Any) -> str:
    if value in ("", None):
        return "blank"
    if field_name == "amount":
        try:
            return f"Rs {format_currency(float(value))}"
        except Exception:
            return str(value)
    if "date" in field_name:
        try:
            return pd.Timestamp(value).strftime(DATE_DISPLAY_FORMAT)
        except Exception:
            return str(value)
    if isinstance(value, Number):
        return f"{value:,.0f}"
    return str(value)


def _build_history_record_label(record: pd.Series, dataset_key: str) -> str:
    if dataset_key == "receivables":
        party = record.get("counterparty", "Receivable") or "Receivable"
        invoice_date = _format_history_value("invoice_date", record.get("invoice_date", ""))
        return f"{party} ({invoice_date})"
    category = record.get("sheet_name", "Other") or "Other"
    vendor = record.get("vendor_name", "").strip() or "Unspecified vendor"
    payment_date = _format_history_value("date", record.get("date", ""))
    return f"{category} / {vendor} ({payment_date})"


def summarize_edit_change(old_df: pd.DataFrame, new_df: pd.DataFrame, dataset_key: str) -> dict[str, Any]:
    config = EDIT_HISTORY_CONFIG[dataset_key]
    old_indexed = _build_history_index(old_df, dataset_key)
    new_indexed = _build_history_index(new_df, dataset_key)

    old_keys = set(old_indexed.index.tolist())
    new_keys = set(new_indexed.index.tolist())
    added_keys = [key for key in new_indexed.index if key not in old_keys]
    removed_keys = [key for key in old_indexed.index if key not in new_keys]
    shared_keys = [key for key in new_indexed.index if key in old_keys]

    compared_columns = [
        column
        for column in config["detail_columns"]
        if column in old_indexed.columns and column in new_indexed.columns
    ]
    modified_rows: list[tuple[str, list[str]]] = []
    for key in shared_keys:
        diff_fields: list[str] = []
        for column in compared_columns:
            old_value = old_indexed.at[key, column]
            new_value = new_indexed.at[key, column]
            if old_value != new_value:
                diff_fields.append(
                    f"{_format_history_field_name(column)}: "
                    f"{_format_history_value(column, old_value)} -> {_format_history_value(column, new_value)}"
                )
        if diff_fields:
            modified_rows.append((key, diff_fields))

    old_amount = pd.to_numeric(old_df.get("amount", pd.Series(dtype=float)), errors="coerce").fillna(0.0).sum()
    new_amount = pd.to_numeric(new_df.get("amount", pd.Series(dtype=float)), errors="coerce").fillna(0.0).sum()
    amount_delta = new_amount - old_amount

    summary_parts: list[str] = []
    if added_keys:
        summary_parts.append(f"{len(added_keys)} added")
    if removed_keys:
        summary_parts.append(f"{len(removed_keys)} removed")
    if modified_rows:
        summary_parts.append(f"{len(modified_rows)} updated")
    if not summary_parts:
        summary_parts.append("No tracked differences")
    if amount_delta:
        direction = "+" if amount_delta > 0 else "-"
        summary_parts.append(f"amount {direction}Rs {format_currency(abs(amount_delta))}")

    details: list[str] = []
    for key, fields in modified_rows[:3]:
        details.append(f"Updated {key.rsplit(' #', 1)[0]}: {'; '.join(fields[:3])}")
    for key in added_keys[:2]:
        record = new_indexed.loc[key]
        details.append(f"Added {_build_history_record_label(record, dataset_key)}")
    for key in removed_keys[:2]:
        record = old_indexed.loc[key]
        details.append(f"Removed {_build_history_record_label(record, dataset_key)}")
    if not details:
        details.append(f"{config['label']} history initialized.")

    return {
        "summary": " | ".join(summary_parts),
        "details": details,
        "count_delta": len(new_df) - len(old_df),
        "amount_delta": amount_delta,
    }


def initialize_edit_history(dataset_key: str, initial_df: pd.DataFrame) -> None:
    snapshot = initial_df.copy()
    st.session_state[_edit_state_key(dataset_key)] = snapshot.copy()
    st.session_state[_edit_history_key(dataset_key)] = [snapshot]
    st.session_state[_edit_history_index_key(dataset_key)] = 0
    st.session_state[_edit_history_log_key(dataset_key)] = []


def commit_edit_history(dataset_key: str, new_df: pd.DataFrame, action_label: str) -> dict[str, Any]:
    state_key = _edit_state_key(dataset_key)
    history_key = _edit_history_key(dataset_key)
    index_key = _edit_history_index_key(dataset_key)
    log_key = _edit_history_log_key(dataset_key)

    current_df = st.session_state[state_key]
    next_df = new_df.copy()
    if edit_frames_equal(current_df, next_df):
        return {"changed": False, "summary": "No changes detected.", "details": []}

    history: list[pd.DataFrame] = st.session_state.get(history_key, [current_df.copy()])
    current_index = st.session_state.get(index_key, len(history) - 1)
    history = history[: current_index + 1]
    change_log: list[dict[str, Any]] = st.session_state.get(log_key, [])[:current_index]

    change_summary = summarize_edit_change(current_df, next_df, dataset_key)
    history.append(next_df.copy())
    change_log.append({
        "timestamp": datetime.now().strftime("%d-%m-%Y %H:%M:%S"),
        "action": action_label,
        "summary": change_summary["summary"],
        "details": change_summary["details"],
    })

    st.session_state[state_key] = next_df
    st.session_state[history_key] = history
    st.session_state[index_key] = len(history) - 1
    st.session_state[log_key] = change_log
    return {"changed": True, **change_summary}


def step_edit_history(dataset_key: str, direction: str) -> bool:
    history_key = _edit_history_key(dataset_key)
    index_key = _edit_history_index_key(dataset_key)
    state_key = _edit_state_key(dataset_key)

    history: list[pd.DataFrame] = st.session_state.get(history_key, [])
    current_index = st.session_state.get(index_key, 0)
    target_index = current_index - 1 if direction == "undo" else current_index + 1
    if target_index < 0 or target_index >= len(history):
        return False

    st.session_state[index_key] = target_index
    st.session_state[state_key] = history[target_index].copy()
    return True


def render_track_changes_panel(dataset_key: str) -> None:
    config = EDIT_HISTORY_CONFIG[dataset_key]
    history: list[pd.DataFrame] = st.session_state.get(_edit_history_key(dataset_key), [])
    history_index = st.session_state.get(_edit_history_index_key(dataset_key), 0)
    change_log: list[dict[str, Any]] = st.session_state.get(_edit_history_log_key(dataset_key), [])

    st.markdown("#### Track Changes")
    st.caption("Undo and redo work on applied changes only.")

    action_col, redo_col, info_col = st.columns([1, 1, 2])
    with action_col:
        undo_label = f"Undo {config['label']}"
        if st.button(undo_label, key=f"undo_{dataset_key}", use_container_width=True, disabled=history_index == 0):
            if step_edit_history(dataset_key, "undo"):
                st.rerun()
    with redo_col:
        redo_label = f"Redo {config['label']}"
        if st.button(
            redo_label,
            key=f"redo_{dataset_key}",
            use_container_width=True,
            disabled=history_index >= max(len(history) - 1, 0),
        ):
            if step_edit_history(dataset_key, "redo"):
                st.rerun()
    with info_col:
        st.caption(
            f"Version {history_index + 1} of {max(len(history), 1)}"
            f" | Undo available: {history_index}"
            f" | Redo available: {max(len(history) - history_index - 1, 0)}"
        )

    if not change_log:
        st.info(f"No committed {config['line_label']} changes yet. Apply edits to start the audit trail.")
        return

    for change_number in range(len(change_log), 0, -1):
        entry = change_log[change_number - 1]
        status = "Active" if change_number <= history_index else "Undone"
        with st.expander(
            f"{status} | {entry['action']} | {entry['timestamp']}",
            expanded=change_number == history_index,
        ):
            st.write(entry["summary"])
            for detail in entry["details"]:
                st.caption(detail)


# ══════════════════════════════════════════════════════════════════════════════
#  HOLIDAY CALENDAR
# ══════════════════════════════════════════════════════════════════════════════

def get_bank_holiday_events(selected_month: pd.Timestamp, holiday_region: str) -> list[dict[str, Any]]:
    month_start = build_month_start(selected_month)
    month_end = month_start + pd.offsets.MonthEnd(1)
    year = month_start.year
    events: dict[pd.Timestamp, list[str]] = {}
    current = month_start
    saturday_count = 0
    while current <= month_end:
        day_events: list[str] = []
        if current.weekday() == 5:
            saturday_count += 1
            if saturday_count == 2:
                day_events.append("Second Saturday")
            elif saturday_count == 4:
                day_events.append("Fourth Saturday")
        public_holiday_name = (
            STATE_BANK_HOLIDAYS.get(holiday_region, {}).get(year, {}).get(current.strftime("%Y-%m-%d"))
        )
        if public_holiday_name:
            day_events.append(public_holiday_name)
        if day_events:
            events[current] = day_events
        current += pd.Timedelta(days=1)
    return [
        {"date": event_date, "name": " / ".join(names)}
        for event_date, names in sorted(events.items(), key=lambda item: item[0])
    ]


def build_week_ranges(selected_month: pd.Timestamp, holiday_region: str) -> list[dict[str, Any]]:
    start = build_month_start(selected_month)
    end = start + pd.offsets.MonthEnd(1)
    holiday_events = get_bank_holiday_events(selected_month, holiday_region)
    weeks = []
    week_start = start
    idx = 1
    while week_start <= end:
        week_end = min(week_start + pd.Timedelta(days=6), end)
        week_holidays = [e for e in holiday_events if week_start <= e["date"] <= week_end]
        weeks.append({
            "key": f"Week {idx}",
            "label": f"Week {idx}: {compact_day_date_label(week_start)} - {compact_day_date_label(week_end)}",
            "start": week_start,
            "end": week_end,
            "holidays": week_holidays,
        })
        week_start = week_end + pd.Timedelta(days=1)
        idx += 1
    return weeks


# ══════════════════════════════════════════════════════════════════════════════
#  FORMULA EVALUATION
# ══════════════════════════════════════════════════════════════════════════════

def classify_aging_bucket(ageing: float) -> str:
    if pd.isna(ageing) or ageing <= 30:
        return "0-30"
    if ageing <= 60:
        return "30-60"
    return "60+"


def is_date_like(value: Any) -> bool:
    return isinstance(value, (pd.Timestamp, date)) or hasattr(value, "year")


def expand_cell_range(range_ref: str) -> list[str]:
    start_ref, end_ref = [part.strip().replace("$", "") for part in range_ref.split(":", 1)]
    start_match = CELL_REF_PATTERN.fullmatch(start_ref)
    end_match = CELL_REF_PATTERN.fullmatch(end_ref)
    if not start_match or not end_match:
        return []
    start_col, start_row = start_match.groups()
    end_col, end_row = end_match.groups()
    refs = []
    for row_idx in range(int(start_row), int(end_row) + 1):
        for col_idx in range(
            column_index_from_string(start_col),
            column_index_from_string(end_col) + 1,
        ):
            refs.append(f"{get_column_letter(col_idx)}{row_idx}")
    return refs


def numeric_or_zero(value: Any) -> float:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return 0.0
    if isinstance(value, (int, float, np.number)):
        return float(value)
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def safe_numeric_eval(expression: str) -> float | None:
    try:
        parsed = ast.parse(expression, mode="eval")
    except SyntaxError:
        return None

    def evaluate_node(node: ast.AST) -> float:
        if isinstance(node, ast.Expression):
            return evaluate_node(node.body)
        if isinstance(node, ast.Constant) and isinstance(node.value, (int, float)):
            return float(node.value)
        if isinstance(node, ast.Num):
            return float(node.n)
        if isinstance(node, ast.BinOp) and isinstance(node.op, ALLOWED_AST_BINARY_OPERATORS):
            left = evaluate_node(node.left)
            right = evaluate_node(node.right)
            if isinstance(node.op, ast.Add):
                return left + right
            if isinstance(node.op, ast.Sub):
                return left - right
            if isinstance(node.op, ast.Mult):
                return left * right
            if isinstance(node.op, ast.Div):
                return left / right
            if isinstance(node.op, ast.Mod):
                return left % right
            return left ** right
        if isinstance(node, ast.UnaryOp) and isinstance(node.op, ALLOWED_AST_UNARY_OPERATORS):
            operand = evaluate_node(node.operand)
            return operand if isinstance(node.op, ast.UAdd) else -operand
        raise ValueError("Unsupported expression")

    try:
        return evaluate_node(parsed)
    except (ValueError, TypeError, ZeroDivisionError, OverflowError):
        return None


def read_workbook_sheets(source: Any) -> dict[str, pd.DataFrame]:
    reset_source_pointer(source)
    workbook_values = load_workbook(source, data_only=True)
    reset_source_pointer(source)
    workbook_formulas = load_workbook(source, data_only=False)
    sheet_frames: dict[str, pd.DataFrame] = {}

    for sheet_name in workbook_values.sheetnames:
        ws_values = workbook_values[sheet_name]
        ws_formulas = workbook_formulas[sheet_name]
        resolved_cache: dict[str, Any] = {}

        def resolve_cell(ref: str, visiting: set[str] | None = None) -> Any:
            cell_ref = ref.replace("$", "")
            if cell_ref in resolved_cache:
                return resolved_cache[cell_ref]
            visiting = visiting or set()
            if cell_ref in visiting:
                return None
            cached_value = ws_values[cell_ref].value
            if cached_value is not None:
                resolved_cache[cell_ref] = cached_value
                return cached_value
            formula_value = ws_formulas[cell_ref].value
            if isinstance(formula_value, str) and formula_value.startswith("="):
                visiting.add(cell_ref)
                resolved = evaluate_formula(formula_value[1:], resolve_cell, visiting)
                visiting.discard(cell_ref)
                resolved_cache[cell_ref] = resolved
                return resolved
            resolved_cache[cell_ref] = formula_value
            return formula_value

        rows: list[list[Any]] = []
        for row_idx in range(1, ws_formulas.max_row + 1):
            row_values = []
            has_content = False
            for col_idx in range(1, ws_formulas.max_column + 1):
                ref = f"{get_column_letter(col_idx)}{row_idx}"
                value = resolve_cell(ref)
                row_values.append(value)
                if value not in (None, "") and not (isinstance(value, float) and pd.isna(value)):
                    has_content = True
            if has_content:
                rows.append(row_values)

        if not rows:
            sheet_frames[sheet_name] = pd.DataFrame()
            continue
        header = []
        for idx, value in enumerate(rows[0], start=1):
            if value in (None, ""):
                header.append(f"Column {idx}")
            else:
                header.append(str(value).strip())
        sheet_frames[sheet_name] = pd.DataFrame(rows[1:], columns=header)

    return sheet_frames


def evaluate_formula(expression: str, resolver: Any, visiting: set[str]) -> Any:
    expr = expression.strip()
    direct_ref = CELL_REF_PATTERN.fullmatch(expr.replace("$", ""))
    if direct_ref:
        return resolver(direct_ref.group(0), visiting)

    date_math = re.fullmatch(r"([A-Z]{1,3}\d+)\s*([+-])\s*(\d+)", expr, flags=re.IGNORECASE)
    if date_math:
        ref, operator, day_count = date_math.groups()
        base_value = resolver(ref, visiting)
        if is_date_like(base_value):
            base_timestamp = pd.Timestamp(base_value)
            delta = pd.Timedelta(days=int(day_count))
            return base_timestamp + delta if operator == "+" else base_timestamp - delta

    def replace_sum(match: re.Match[str]) -> str:
        total = 0.0
        parts = [part.strip() for part in match.group(1).split(",") if part.strip()]
        for part in parts:
            if ":" in part:
                refs = expand_cell_range(part)
                total += sum(numeric_or_zero(resolver(r, visiting)) for r in refs)
            elif CELL_REF_PATTERN.fullmatch(part.replace("$", "")):
                total += numeric_or_zero(resolver(part, visiting))
            else:
                total += numeric_or_zero(evaluate_formula(part, resolver, visiting))
        return str(total)

    expr = re.sub(r"(?i)SUM\(([^)]+)\)", replace_sum, expr)
    expr = re.sub(r"(\d+(?:\.\d+)?)%", lambda m: f"({m.group(1)}/100)", expr)

    def replace_cell(match: re.Match[str]) -> str:
        ref = match.group(0)
        value = resolver(ref, visiting)
        if is_date_like(value):
            return "0"
        return str(numeric_or_zero(value))

    expr = CELL_REF_PATTERN.sub(replace_cell, expr)
    return safe_numeric_eval(expr)


# ══════════════════════════════════════════════════════════════════════════════
#  DATA NORMALISATION
# ══════════════════════════════════════════════════════════════════════════════

def normalize_receivables(
    df: pd.DataFrame,
    sheet_name: str = "Receivables",
    actual_mode: bool = False,
    column_map: dict[str, str | None] | None = None,
) -> pd.DataFrame:
    empty_cols = [
        "invoice_date", "cash_date", "tentative_collection_date", "due_date",
        "billing_type", "billing_bucket", "source_sheet", "description", "counterparty",
        "amount", "ageing", "aging_bucket",
    ]
    if df.empty:
        return pd.DataFrame(columns=empty_cols)

    column_map = column_map or detect_receivables_columns(list(df.columns))
    date_col = column_map.get("date")
    amount_col = column_map.get("amount")
    client_col = column_map.get("client")
    ref_col = column_map.get("reference")
    ageing_col = column_map.get("ageing")
    tentative_date_col = column_map.get("tentative_date")
    due_date_col = column_map.get("due_date")
    billing_type_col = column_map.get("billing_type")

    if amount_col is None:
        return pd.DataFrame(columns=empty_cols)

    invoice_date = to_datetime(df[date_col]) if date_col else pd.Series(pd.NaT, index=df.index)
    amount = to_numeric(df[amount_col])
    counterparty = df[client_col].fillna("Receivable") if client_col else pd.Series("Receivable", index=df.index)
    reference = df[ref_col].fillna("") if ref_col else pd.Series("", index=df.index)
    description = np.where(
        reference.astype(str).str.strip() != "",
        counterparty.astype(str) + " | " + reference.astype(str),
        counterparty.astype(str),
    )
    ageing = to_numeric(df[ageing_col]) if ageing_col else pd.Series(0, index=df.index, dtype=float)
    aging_bucket = ageing.apply(classify_aging_bucket)
    cash_date = invoice_date if actual_mode else pd.Series(pd.NaT, index=df.index)
    tentative_collection_date = to_datetime(df[tentative_date_col]) if tentative_date_col else pd.Series(pd.NaT, index=df.index)
    due_date = to_datetime(df[due_date_col]) if due_date_col else pd.Series(pd.NaT, index=df.index)
    billing_type = df[billing_type_col].fillna("").astype(str).str.strip() if billing_type_col else pd.Series("", index=df.index, dtype="object")
    billing_bucket = billing_type.map(normalize_billing_bucket)

    normalized = pd.DataFrame({
        "invoice_date": invoice_date,
        "cash_date": cash_date,
        "tentative_collection_date": tentative_collection_date,
        "due_date": due_date,
        "billing_type": billing_type,
        "billing_bucket": billing_bucket,
        "source_sheet": sheet_name,
        "description": description,
        "counterparty": counterparty,
        "amount": amount,
        "ageing": ageing,
        "aging_bucket": aging_bucket,
    })
    normalized = normalized.loc[normalized["amount"] != 0].copy()
    return normalized.reset_index(drop=True)


def normalize_outflow_sheet(
    df: pd.DataFrame,
    sheet_name: str,
    actual_mode: bool = False,
    column_map: dict[str, str | None] | None = None,
) -> pd.DataFrame:
    empty_cols = ["sheet_name", "date", "description", "vendor_name", "amount"]
    if df.empty:
        return pd.DataFrame(columns=empty_cols)

    column_map = column_map or detect_outflow_columns(list(df.columns))
    amount_col = column_map.get("amount")
    if amount_col is None:
        return pd.DataFrame(columns=empty_cols)

    if actual_mode:
        actual_date_col = column_map.get("actual_date")
        planned_date_col = column_map.get("planned_date")
        if actual_date_col and planned_date_col and actual_date_col != planned_date_col:
            date_values = to_datetime(df[actual_date_col]).fillna(to_datetime(df[planned_date_col]))
        elif actual_date_col:
            date_values = to_datetime(df[actual_date_col])
        elif planned_date_col:
            date_values = to_datetime(df[planned_date_col])
        else:
            date_values = pd.Series(pd.NaT, index=df.index)
    else:
        date_col = column_map.get("planned_date") or column_map.get("actual_date")
        date_values = to_datetime(df[date_col]) if date_col else pd.Series(pd.NaT, index=df.index)

    vendor_col = column_map.get("vendor")
    description_col = column_map.get("description")

    if date_values.isna().all():
        return pd.DataFrame(columns=empty_cols)

    vendor_name = build_summary_label(df, vendor_col, description_col, sheet_name)
    description = build_summary_label(df, description_col, vendor_col, sheet_name)

    normalized = pd.DataFrame({
        "sheet_name": sheet_name,
        "date": date_values,
        "description": description,
        "vendor_name": vendor_name,
        "amount": to_numeric(df[amount_col]),
    })
    normalized = normalized.loc[normalized["date"].notna() & (normalized["amount"] != 0)].copy()
    return normalized.reset_index(drop=True)


def load_data(source: Any, actual_mode: bool = False) -> dict[str, Any]:
    receivable_frames: list[pd.DataFrame] = []
    inflow_sheet_order: list[str] = []
    outflow_frames: list[pd.DataFrame] = []
    outflow_sheet_order: list[str] = []
    workbook_sheets = read_workbook_sheets(source)
    workbook_layout: dict[str, Any] = {
        "sheet_order": list(workbook_sheets.keys()),
        "sheet_meta": {},
        "inflow_sheet_names": [],
        "outflow_sheet_names": [],
    }

    for sheet_name, raw in workbook_sheets.items():
        raw = raw.dropna(how="all")
        schema = classify_sheet_schema(sheet_name, raw) if not raw.empty else {"role": "ignored", "column_map": {}, "columns": list(raw.columns)}
        workbook_layout["sheet_meta"][sheet_name] = schema
        if raw.empty:
            continue
        if schema["role"] == "receivables":
            normalized = normalize_receivables(
                raw,
                sheet_name=sheet_name,
                actual_mode=actual_mode,
                column_map=schema["column_map"],
            )
            if not normalized.empty:
                receivable_frames.append(normalized)
                inflow_sheet_order.append(sheet_name)
                workbook_layout["inflow_sheet_names"].append(sheet_name)
        elif schema["role"] == "outflow":
            normalized = normalize_outflow_sheet(raw, sheet_name, actual_mode=actual_mode, column_map=schema["column_map"])
            if not normalized.empty:
                outflow_frames.append(normalized)
                outflow_sheet_order.append(sheet_name)
                workbook_layout["outflow_sheet_names"].append(sheet_name)

    receivables = (
        pd.concat(receivable_frames, ignore_index=True)
        if receivable_frames
        else pd.DataFrame(columns=[
            "invoice_date", "cash_date", "tentative_collection_date", "due_date",
            "billing_type", "billing_bucket", "source_sheet", "description", "counterparty",
            "amount", "ageing", "aging_bucket",
        ])
    )
    outflows = (
        pd.concat(outflow_frames, ignore_index=True)
        if outflow_frames
        else pd.DataFrame(columns=["sheet_name", "date", "description", "vendor_name", "amount"])
    )
    inflow_sheet_order = order_sheet_names(inflow_sheet_order)
    outflow_sheet_order = order_outflow_sheets(outflow_sheet_order)
    month_candidates = []
    if not outflows.empty:
        month_candidates.extend(outflows["date"].dropna().tolist())
    if not receivables.empty:
        for date_column in ["tentative_collection_date", "due_date", "invoice_date"]:
            if date_column in receivables.columns:
                month_candidates.extend(receivables[date_column].dropna().tolist())
    if actual_mode and not receivables.empty:
        month_candidates.extend(receivables["cash_date"].dropna().tolist())

    available_months = sorted({build_month_start(item) for item in month_candidates})
    return {
        "receivables": receivables,
        "inflow_sheet_order": inflow_sheet_order,
        "outflows": outflows,
        "outflow_sheet_order": outflow_sheet_order,
        "available_months": available_months,
        "workbook_layout": workbook_layout,
    }


# ══════════════════════════════════════════════════════════════════════════════
#  COLLECTION PROBABILITY ENGINE  (now configurable)
# ══════════════════════════════════════════════════════════════════════════════

def derive_collection_probabilities(
    receivables_df: pd.DataFrame,
    anchor_date: pd.Timestamp | None = None,
    *,
    base_probs: dict[str, float] | None = None,
    worst_probs: dict[str, float] | None = None,
    tds_rate: float = 0.0,
) -> pd.DataFrame:
    """
    Attach collection probability columns to the receivables dataframe.

    Parameters
    ----------
    base_probs / worst_probs : override the module-level defaults per bucket.
    tds_rate : fraction deducted at source (e.g. 0.02 for 2% TDS).
               Applied AFTER probability so expected cash = gross × prob × (1 − tds).
    """
    if receivables_df.empty:
        return receivables_df.copy()

    eff_base = base_probs or DEFAULT_BASE_COLLECTION_PROBABILITIES
    eff_worst = worst_probs or DEFAULT_WORST_COLLECTION_PROBABILITIES
    net_factor = 1.0 - max(0.0, min(tds_rate, 1.0))

    anchor = build_month_start(anchor_date or TODAY)
    derived = receivables_df.copy()
    derived["base_probability"] = derived["aging_bucket"].map(eff_base).fillna(0.0)
    derived["best_probability"] = derived["base_probability"]
    derived["worst_probability"] = derived["aging_bucket"].map(eff_worst).fillna(0.0)
    derived["expected_base"] = derived["amount"] * derived["base_probability"] * net_factor
    derived["expected_best"] = derived["amount"] * derived["best_probability"] * net_factor
    derived["expected_worst"] = derived["amount"] * derived["worst_probability"] * net_factor

    fallback_dates = derived["aging_bucket"].map(
        lambda bucket: anchor + pd.Timedelta(days=COLLECTION_DATE_OFFSETS[bucket])
    )
    if "tentative_collection_date" in derived.columns:
        derived["forecast_collection_date"] = derived["tentative_collection_date"].fillna(fallback_dates)
    else:
        derived["forecast_collection_date"] = fallback_dates
    return derived


def actual_receivables_look_plan_aligned(receivables_df: pd.DataFrame) -> bool:
    required_columns = {"invoice_date", "cash_date", "tentative_collection_date", "ageing"}
    if receivables_df.empty or not required_columns.issubset(receivables_df.columns):
        return False
    if not receivables_df["tentative_collection_date"].notna().any():
        return False
    comparable = receivables_df.loc[
        receivables_df["cash_date"].notna() & receivables_df["invoice_date"].notna(),
        ["cash_date", "invoice_date"],
    ]
    if comparable.empty:
        return False
    return (comparable["cash_date"] == comparable["invoice_date"]).all()


def build_actual_receivables_view(
    actual_data: dict[str, Any],
    anchor_date: pd.Timestamp,
    *,
    base_probs: dict[str, float] | None = None,
    worst_probs: dict[str, float] | None = None,
    tds_rate: float = 0.0,
) -> pd.DataFrame:
    receivables = actual_data.get("receivables", pd.DataFrame()).copy()
    if receivables.empty:
        return receivables
    if actual_receivables_look_plan_aligned(receivables):
        aligned = derive_collection_probabilities(
            receivables,
            anchor_date=anchor_date,
            base_probs=base_probs,
            worst_probs=worst_probs,
            tds_rate=tds_rate,
        )
        aligned["comparison_date"] = aligned["forecast_collection_date"]
        aligned["comparison_amount"] = aligned["expected_base"]
        return aligned
    receivables["comparison_date"] = receivables["cash_date"]
    receivables["comparison_amount"] = receivables["amount"]
    return receivables


def split_weekly_receivable_buckets(
    receipts_df: pd.DataFrame,
    start: pd.Timestamp,
    end: pd.Timestamp,
    *,
    value_column: str,
) -> dict[str, float]:
    if receipts_df.empty:
        return {
            "advance": 0.0,
            "receivables": 0.0,
            "due_within_week": 0.0,
            "overdue": 0.0,
            "total": 0.0,
        }

    bucket_masks = classify_weekly_receivable_masks(receipts_df, start, end)
    advance_mask = bucket_masks["advance"]
    receivable_mask = bucket_masks["receivables"]
    due_within_week_mask = bucket_masks["due_within_week"]
    overdue_mask = bucket_masks["overdue"]

    return {
        "advance": receipts_df.loc[advance_mask, value_column].sum(),
        "receivables": receipts_df.loc[receivable_mask, value_column].sum(),
        "due_within_week": receipts_df.loc[due_within_week_mask, value_column].sum(),
        "overdue": receipts_df.loc[overdue_mask, value_column].sum(),
        "total": receipts_df[value_column].sum(),
    }


def classify_weekly_receivable_masks(
    receipts_df: pd.DataFrame,
    start: pd.Timestamp,
    end: pd.Timestamp,
) -> dict[str, pd.Series]:
    billing_bucket = receipts_df.get("billing_bucket", pd.Series("Receivables", index=receipts_df.index)).fillna("Receivables")
    due_dates = (
        to_datetime(receipts_df["due_date"])
        if "due_date" in receipts_df.columns
        else pd.Series(pd.NaT, index=receipts_df.index)
    )
    # Weekly collections are already filtered to the week by tentative / forecast collection date.
    # Within that bucket:
    # - Advance = invoices not yet due as of the week end, plus rows explicitly tagged Advance
    # - Receivables = invoices already due or becoming due within the week
    explicit_advance_mask = billing_bucket == "Advance"
    future_due_mask = due_dates.notna() & (due_dates > end)
    advance_mask = explicit_advance_mask | future_due_mask
    receivable_mask = ~advance_mask
    overdue_mask = receivable_mask & due_dates.notna() & (due_dates < start)
    due_within_week_mask = receivable_mask & due_dates.between(start, end, inclusive="both")
    # If due date is blank on a non-advance row, keep it under receivables and default it to due within week.
    unknown_due_mask = receivable_mask & due_dates.isna()
    due_within_week_mask = due_within_week_mask | unknown_due_mask

    return {
        "advance": advance_mask,
        "receivables": receivable_mask,
        "due_within_week": due_within_week_mask,
        "overdue": overdue_mask,
    }


# ══════════════════════════════════════════════════════════════════════════════
#  FILTERING & HORIZON HELPERS
# ══════════════════════════════════════════════════════════════════════════════

def filter_between(df: pd.DataFrame, date_column: str, start: pd.Timestamp, end: pd.Timestamp) -> pd.DataFrame:
    if df.empty:
        return df.copy()
    return df.loc[df[date_column].between(start, end, inclusive="both")].copy()


def build_month_horizon(selected_month: pd.Timestamp, periods: int = 3) -> list[pd.Timestamp]:
    start = build_month_start(selected_month)
    return [start + pd.DateOffset(months=offset) for offset in range(periods)]


# ══════════════════════════════════════════════════════════════════════════════
#  MONTHLY ROLL-FORWARD
# ══════════════════════════════════════════════════════════════════════════════

def compute_monthly_rollforward(
    receivables_df: pd.DataFrame,
    outflows_df: pd.DataFrame,
    months: list[pd.Timestamp],
    opening_balance: float,
    scenario_column: str = "expected_base",
    period_start_overrides: dict[pd.Timestamp, pd.Timestamp] | None = None,
) -> pd.DataFrame:
    rows = []
    running_open = opening_balance
    for month_start in months:
        effective_start = (period_start_overrides or {}).get(month_start, month_start)
        month_end = month_start + pd.offsets.MonthEnd(1)
        inflows = (
            filter_between(receivables_df, "forecast_collection_date", effective_start, month_end)[scenario_column].sum()
            if not receivables_df.empty
            else 0.0
        )
        outflows = (
            filter_between(outflows_df, "date", effective_start, month_end)["amount"].sum()
            if not outflows_df.empty
            else 0.0
        )
        closing = running_open + inflows - outflows
        rows.append({
            "Month": month_label(month_start),
            "Opening Cash": running_open,
            "Inflows": inflows,
            "Outflows": outflows,
            "Closing Cash": closing,
            "month_start": month_start,
        })
        running_open = closing
    return pd.DataFrame(rows)


def compute_od_cash_movement(
    opening_cash: float,
    opening_od_outstanding: float,
    collections: float,
    outflows: float,
    od_sanctioned_limit: float,
    *,
    cash_buffer: float = DEFAULT_MIN_CASH_THRESHOLD,
) -> dict[str, float]:
    cash_before_od = opening_cash + collections - outflows
    draw_capacity = (
        max(od_sanctioned_limit - opening_od_outstanding, 0.0)
        if od_sanctioned_limit > 0
        else np.inf
    )

    od_movement = 0.0
    ending_od_outstanding = opening_od_outstanding
    ending_cash = cash_before_od

    if cash_before_od < 0:
        required_draw = -cash_before_od
        draw_amount = min(required_draw, draw_capacity)
        od_movement = draw_amount
        ending_od_outstanding = opening_od_outstanding + draw_amount
        ending_cash = cash_before_od + draw_amount
    elif opening_od_outstanding > 0 and cash_before_od > cash_buffer:
        repay_amount = min(cash_before_od - cash_buffer, opening_od_outstanding)
        od_movement = -repay_amount
        ending_od_outstanding = opening_od_outstanding - repay_amount
        ending_cash = cash_before_od - repay_amount

    return {
        "opening_cash": opening_cash,
        "opening_od_outstanding": opening_od_outstanding,
        "cash_before_od": cash_before_od,
        "od_movement": od_movement,
        "ending_od_outstanding": ending_od_outstanding,
        "ending_cash": ending_cash,
        "available_headroom": (
            max(od_sanctioned_limit - ending_od_outstanding, 0.0)
            if od_sanctioned_limit > 0
            else 0.0
        ),
    }


def build_monthly_detail_matrix(
    df: pd.DataFrame,
    date_column: str,
    value_column: str,
    line_item_column: str,
    months: list[pd.Timestamp],
    line_item_label: str,
    period_start_overrides: dict[pd.Timestamp, pd.Timestamp] | None = None,
) -> pd.DataFrame:
    month_labels = [month_label(month) for month in months]
    if df.empty:
        return pd.DataFrame(columns=["Line Item"] + month_labels)

    detail = df.copy()
    detail["forecast_month"] = pd.to_datetime(detail[date_column], errors="coerce").dt.to_period("M").dt.to_timestamp()
    detail = detail.loc[detail["forecast_month"].isin(months)].copy()
    detail["period_start"] = detail["forecast_month"].map(
        lambda month: (period_start_overrides or {}).get(month, month)
    )
    detail = detail.loc[pd.to_datetime(detail[date_column], errors="coerce") >= detail["period_start"]].copy()
    if detail.empty:
        return pd.DataFrame(columns=["Line Item"] + month_labels)

    pivot = (
        detail.pivot_table(
            index=line_item_column,
            columns="forecast_month",
            values=value_column,
            aggfunc="sum",
            fill_value=0.0,
        )
        .reindex(columns=months, fill_value=0.0)
        .sort_index()
    )
    pivot.columns = month_labels
    pivot = pivot.reset_index().rename(columns={line_item_column: "Line Item"})
    total_row = {"Line Item": line_item_label}
    for lbl in month_labels:
        total_row[lbl] = pivot[lbl].sum()
    return pd.concat([pd.DataFrame([total_row]), pivot], ignore_index=True)


# ══════════════════════════════════════════════════════════════════════════════
#  WEEK DETAIL COMPUTATION
# ══════════════════════════════════════════════════════════════════════════════

def build_week_detail(
    start: pd.Timestamp,
    end: pd.Timestamp,
    plan_receivables: pd.DataFrame,
    plan_outflows: pd.DataFrame,
    outflow_sheet_order: list[str],
    actual_data: dict[str, Any] | None = None,
    effective_start: pd.Timestamp | None = None,
    actual_anchor_date: pd.Timestamp | None = None,
    *,
    base_probs: dict[str, float] | None = None,
    worst_probs: dict[str, float] | None = None,
    tds_rate: float = 0.0,
) -> dict[str, Any]:
    active_start = max(start, effective_start) if effective_start is not None else start
    if active_start > end:
        week_plan_receipts = pd.DataFrame(columns=plan_receivables.columns)
        week_plan_outflows = pd.DataFrame(columns=plan_outflows.columns)
    else:
        week_plan_receipts = filter_between(plan_receivables, "forecast_collection_date", active_start, end)
        week_plan_outflows = filter_between(plan_outflows, "date", active_start, end)

    plan_receivable_split = split_weekly_receivable_buckets(
        week_plan_receipts,
        active_start,
        end,
        value_column="expected_base",
    )
    plan_advance_collections = plan_receivable_split["advance"]
    plan_receivable_collections = plan_receivable_split["receivables"]
    plan_due_within_week_collections = plan_receivable_split["due_within_week"]
    plan_overdue_collections = plan_receivable_split["overdue"]
    plan_collections = plan_receivable_split["total"]
    plan_by_category = (
        week_plan_outflows.groupby("sheet_name", as_index=True)["amount"]
        .sum()
        .reindex(outflow_sheet_order, fill_value=0.0)
        if not week_plan_outflows.empty
        else pd.Series(0.0, index=outflow_sheet_order)
    )
    total_outflows = plan_by_category.sum()
    net_plan = plan_collections - total_outflows

    actual_collections = 0.0
    actual_by_category = pd.Series(0.0, index=outflow_sheet_order)
    actual_net = None
    actual_week_receipts = pd.DataFrame()
    actual_week_outflows = pd.DataFrame()
    actual_advance_collections = 0.0
    actual_receivable_collections = 0.0
    actual_due_within_week_collections = 0.0
    actual_overdue_collections = 0.0

    if actual_data:
        actual_receivables = build_actual_receivables_view(
            actual_data,
            anchor_date=actual_anchor_date or active_start,
            base_probs=base_probs,
            worst_probs=worst_probs,
            tds_rate=tds_rate,
        )
        actual_outflows_df = actual_data["outflows"]
        if active_start <= end:
            actual_week_receipts = filter_between(actual_receivables, "comparison_date", active_start, end)
            actual_week_outflows = filter_between(actual_outflows_df, "date", active_start, end)
        actual_receivable_split = split_weekly_receivable_buckets(
            actual_week_receipts,
            active_start,
            end,
            value_column="comparison_amount",
        )
        actual_collections = actual_receivable_split["total"]
        actual_advance_collections = actual_receivable_split["advance"]
        actual_receivable_collections = actual_receivable_split["receivables"]
        actual_due_within_week_collections = actual_receivable_split["due_within_week"]
        actual_overdue_collections = actual_receivable_split["overdue"]
        actual_by_category = (
            actual_week_outflows.groupby("sheet_name", as_index=True)["amount"]
            .sum()
            .reindex(outflow_sheet_order, fill_value=0.0)
            if not actual_week_outflows.empty
            else pd.Series(0.0, index=outflow_sheet_order)
        )
        actual_net = actual_collections - actual_by_category.sum()

    return {
        "start": start,
        "end": end,
        "effective_start": active_start,
        "plan_receipts_detail": week_plan_receipts,
        "plan_outflows_detail": week_plan_outflows,
        "plan_advance_collections": plan_advance_collections,
        "plan_receivable_collections": plan_receivable_collections,
        "plan_due_within_week_collections": plan_due_within_week_collections,
        "plan_overdue_collections": plan_overdue_collections,
        "plan_collections": plan_collections,
        "plan_by_category": plan_by_category,
        "plan_total_outflows": total_outflows,
        "plan_net": net_plan,
        "actual_receipts_detail": actual_week_receipts,
        "actual_outflows_detail": actual_week_outflows,
        "actual_advance_collections": actual_advance_collections,
        "actual_receivable_collections": actual_receivable_collections,
        "actual_due_within_week_collections": actual_due_within_week_collections,
        "actual_overdue_collections": actual_overdue_collections,
        "actual_collections": actual_collections,
        "actual_by_category": actual_by_category,
        "actual_net": actual_net,
    }


# ══════════════════════════════════════════════════════════════════════════════
#  WEEKLY CASH FLOW COMPUTATION
# ══════════════════════════════════════════════════════════════════════════════

def compute_weekly_cashflow(
    data: dict[str, Any],
    selected_month: pd.Timestamp,
    opening_balance: float,
    holiday_region: str,
    balance_as_of_date: pd.Timestamp | None = None,
    actual_data: dict[str, Any] | None = None,
    *,
    base_probs: dict[str, float] | None = None,
    worst_probs: dict[str, float] | None = None,
    tds_rate: float = 0.0,
    od_sanctioned_limit: float = 0.0,
    od_interest_rate_pa: float = 0.0,
    opening_od_utilization: float = 0.0,
    od_buffer_balance: float = DEFAULT_MIN_CASH_THRESHOLD,
) -> dict[str, Any]:
    month_start = build_month_start(selected_month)
    month_end = build_month_end(selected_month)
    balance_date = pd.Timestamp(balance_as_of_date or month_start).normalize()
    balance_date = min(max(balance_date, month_start), month_end)
    period_start_overrides = {month_start: balance_date}

    plan_receivables = derive_collection_probabilities(
        data["receivables"],
        anchor_date=balance_date,
        base_probs=base_probs,
        worst_probs=worst_probs,
        tds_rate=tds_rate,
    )
    plan_outflows = data["outflows"].copy()
    outflow_sheet_order = data["outflow_sheet_order"]
    weeks = build_week_ranges(month_start, holiday_region)

    horizon_months = build_month_horizon(month_start, periods=3)
    monthly_forecast = compute_monthly_rollforward(
        plan_receivables,
        plan_outflows,
        horizon_months,
        opening_balance,
        scenario_column="expected_base",
        period_start_overrides=period_start_overrides,
    )
    monthly_inflow_detail = build_monthly_detail_matrix(
        plan_receivables,
        date_column="forecast_collection_date",
        value_column="expected_base",
        line_item_column="counterparty",
        months=horizon_months,
        line_item_label="Collections",
        period_start_overrides=period_start_overrides,
    )
    monthly_outflow_detail = build_monthly_detail_matrix(
        plan_outflows,
        date_column="date",
        value_column="amount",
        line_item_column="sheet_name",
        months=horizon_months,
        line_item_label="Cash Outflows",
        period_start_overrides=period_start_overrides,
    )

    running_plan = opening_balance
    running_plan_od = max(opening_od_utilization, 0.0)
    running_actual = opening_balance if actual_data else None
    running_actual_od = max(opening_od_utilization, 0.0) if actual_data else None
    week_meta: list[dict[str, Any]] = []

    for week in weeks:
        week_detail = build_week_detail(
            week["start"],
            week["end"],
            plan_receivables,
            plan_outflows,
            outflow_sheet_order,
            actual_data=actual_data,
            effective_start=balance_date,
            actual_anchor_date=balance_date,
            base_probs=base_probs,
            worst_probs=worst_probs,
            tds_rate=tds_rate,
        )

        # ── OD utilisation & interest for this week ───────────────────────
        days_in_week = max((week["end"] - week["start"]).days + 1, 1)

        if week_detail["effective_start"] > week_detail["end"]:
            week_detail["opening_plan"] = None
            week_detail["ending_plan"] = None
            week_detail["opening_plan_od_outstanding"] = None
            week_detail["plan_od_utilization"] = None
            week_detail["ending_plan_od_outstanding"] = None
            week_detail["plan_od_headroom"] = None
            week_detail["plan_total_inflows"] = None
            week_detail["plan_surplus_deficit"] = None
            week_detail["plan_od_interest"] = None
        else:
            week_detail["opening_plan"] = running_plan
            week_detail["opening_plan_od_outstanding"] = running_plan_od
            od_result = compute_od_cash_movement(
                running_plan,
                running_plan_od,
                week_detail["plan_collections"],
                week_detail["plan_total_outflows"],
                od_sanctioned_limit,
                cash_buffer=od_buffer_balance,
            )
            week_detail["plan_od_utilization"] = od_result["od_movement"]
            week_detail["ending_plan_od_outstanding"] = od_result["ending_od_outstanding"]
            week_detail["plan_od_headroom"] = od_result["available_headroom"]
            week_detail["plan_total_inflows"] = week_detail["plan_collections"]
            week_detail["plan_net"] = (
                week_detail["plan_total_inflows"] - week_detail["plan_total_outflows"]
            )
            week_detail["plan_surplus_deficit"] = od_result["cash_before_od"]
            week_detail["ending_plan"] = od_result["ending_cash"]
            # Estimated OD interest for this week
            week_detail["plan_od_interest"] = (
                ((running_plan_od + od_result["ending_od_outstanding"]) / 2.0)
                * (od_interest_rate_pa / 100.0) / 365.0 * days_in_week
                if od_interest_rate_pa > 0 and (running_plan_od > 0 or od_result["ending_od_outstanding"] > 0)
                else 0.0
            )
            running_plan = week_detail["ending_plan"]
            running_plan_od = od_result["ending_od_outstanding"]

        if actual_data:
            if week_detail["effective_start"] > week_detail["end"]:
                week_detail["opening_actual"] = None
                week_detail["ending_actual"] = None
                week_detail["opening_actual_od_outstanding"] = None
                week_detail["actual_od_utilization"] = None
                week_detail["ending_actual_od_outstanding"] = None
                week_detail["actual_od_headroom"] = None
                week_detail["actual_total_inflows"] = None
                week_detail["unexpected"] = None
                week_detail["actual_od_interest"] = None
            else:
                week_detail["opening_actual"] = running_actual
                week_detail["opening_actual_od_outstanding"] = running_actual_od
                actual_outflow_total = week_detail["actual_by_category"].sum()
                actual_od_result = compute_od_cash_movement(
                    running_actual,
                    running_actual_od or 0.0,
                    week_detail["actual_collections"],
                    actual_outflow_total,
                    od_sanctioned_limit,
                    cash_buffer=od_buffer_balance,
                )
                week_detail["actual_od_utilization"] = actual_od_result["od_movement"]
                week_detail["ending_actual_od_outstanding"] = actual_od_result["ending_od_outstanding"]
                week_detail["actual_od_headroom"] = actual_od_result["available_headroom"]
                week_detail["actual_total_inflows"] = week_detail["actual_collections"]
                actual_net = week_detail["actual_total_inflows"] - actual_outflow_total
                week_detail["actual_net"] = actual_net
                week_detail["actual_surplus_deficit"] = actual_od_result["cash_before_od"]
                week_detail["ending_actual"] = actual_od_result["ending_cash"]
                running_actual = week_detail["ending_actual"]
                running_actual_od = actual_od_result["ending_od_outstanding"]
                week_detail["unexpected"] = actual_net - (week_detail["plan_net"] or 0.0)
                week_detail["actual_od_interest"] = (
                    (((week_detail["opening_actual_od_outstanding"] or 0.0) + actual_od_result["ending_od_outstanding"]) / 2.0)
                    * (od_interest_rate_pa / 100.0) / 365.0 * days_in_week
                    if od_interest_rate_pa > 0 and (
                        (week_detail["opening_actual_od_outstanding"] or 0.0) > 0
                        or actual_od_result["ending_od_outstanding"] > 0
                    )
                    else 0.0
                )
        else:
            week_detail["opening_actual"] = None
            week_detail["ending_actual"] = None
            week_detail["opening_actual_od_outstanding"] = None
            week_detail["actual_od_utilization"] = None
            week_detail["ending_actual_od_outstanding"] = None
            week_detail["actual_od_headroom"] = None
            week_detail["actual_total_inflows"] = None
            week_detail["actual_surplus_deficit"] = None
            week_detail["unexpected"] = (
                None if week_detail["effective_start"] > week_detail["end"] else 0.0
            )
            week_detail["actual_od_interest"] = None

        week_meta.append(week_detail | {
            "key": week["key"],
            "label": week["label"],
            "holidays": week.get("holidays", []),
        })

    # ── Build table rows ──────────────────────────────────────────────────────
    line_items = [
        "Inflows",
        "   Advance",
        "   Receivables",
        "      Due Within Week",
        "      Overdue",
        "Total Inflows",
        "Cash Outflows",
    ]
    line_items.extend([f"   {sheet}" for sheet in outflow_sheet_order])
    line_items.extend([
        "Total Outflows",
        "Net Movement of Cash",
        "Opening Balance incl. OD Opening O/S",
        "Closing Balance",
    ])

    week_values: dict[str, dict[str, Any]] = {item: {} for item in line_items}

    for wd in week_meta:
        key = wd["key"]
        if wd["effective_start"] > wd["end"]:
            for item in line_items:
                week_values[item][key] = None
            continue
        week_values["Inflows"][key] = None
        week_values["   Advance"][key] = wd["plan_advance_collections"]
        week_values["   Receivables"][key] = wd["plan_receivable_collections"]
        week_values["      Due Within Week"][key] = wd["plan_due_within_week_collections"]
        week_values["      Overdue"][key] = wd["plan_overdue_collections"]
        week_values["Total Inflows"][key] = wd["plan_total_inflows"]
        week_values["Cash Outflows"][key] = None
        for sheet in outflow_sheet_order:
            week_values[f"   {sheet}"][key] = wd["plan_by_category"].get(sheet, 0.0)
        week_values["Total Outflows"][key] = wd["plan_total_outflows"]
        week_values["Net Movement of Cash"][key] = wd["plan_net"]
        week_values["Opening Balance incl. OD Opening O/S"][key] = (
            (wd["opening_plan"] or 0.0) - (wd["opening_plan_od_outstanding"] or 0.0)
            if wd["opening_plan"] is not None
            else None
        )
        week_values["Closing Balance"][key] = (
            (wd["ending_plan"] or 0.0) - (wd["ending_plan_od_outstanding"] or 0.0)
            if wd["ending_plan"] is not None
            else None
        )

    active_weeks = [wd for wd in week_meta if wd["effective_start"] <= wd["end"]]
    monthly_plan_advance = sum(wd["plan_advance_collections"] for wd in active_weeks)
    monthly_plan_receivables = sum(wd["plan_receivable_collections"] for wd in active_weeks)
    monthly_plan_due_within_week = sum(wd["plan_due_within_week_collections"] for wd in active_weeks)
    monthly_plan_overdue = sum(wd["plan_overdue_collections"] for wd in active_weeks)
    monthly_plan_collections = sum(wd["plan_collections"] for wd in active_weeks)
    monthly_plan_od_utilization = sum(wd["plan_od_utilization"] or 0.0 for wd in active_weeks)
    monthly_plan_total_inflows = sum(wd["plan_total_inflows"] or 0.0 for wd in active_weeks)
    monthly_plan_by_category = (
        pd.DataFrame([wd["plan_by_category"] for wd in active_weeks])
        .fillna(0.0)
        .sum()
        .reindex(outflow_sheet_order, fill_value=0.0)
        if active_weeks
        else pd.Series(0.0, index=outflow_sheet_order)
    )
    monthly_plan_outflows = sum(wd["plan_total_outflows"] for wd in active_weeks)
    monthly_plan_net = sum(wd["plan_net"] for wd in active_weeks)
    monthly_plan_opening_incl_od = opening_balance - max(opening_od_utilization, 0.0)
    monthly_plan_ending = active_weeks[-1]["ending_plan"] if active_weeks else opening_balance
    monthly_plan_od_interest = sum(wd.get("plan_od_interest") or 0.0 for wd in active_weeks)
    monthly_plan_ending_od = active_weeks[-1]["ending_plan_od_outstanding"] if active_weeks else running_plan_od
    monthly_plan_closing = monthly_plan_ending - monthly_plan_ending_od
    monthly_plan_headroom = (
        max(od_sanctioned_limit - monthly_plan_ending_od, 0.0)
        if od_sanctioned_limit > 0
        else 0.0
    )

    # ── Actuals monthly aggregation ───────────────────────────────────────────
    actual_month_collections = actual_month_advance = actual_month_receivables = None
    actual_month_due_within_week = actual_month_overdue = None
    actual_month_od_utilization = actual_month_total_inflows = None
    actual_month_by_category = pd.Series(dtype=float)
    actual_month_outflows = actual_month_net = actual_month_ending = None
    actual_month_ending_od = None
    actual_month_closing = None

    if actual_data:
        actual_month_advance = sum((wd.get("actual_advance_collections") or 0.0) for wd in active_weeks)
        actual_month_receivables = sum((wd.get("actual_receivable_collections") or 0.0) for wd in active_weeks)
        actual_month_due_within_week = sum((wd.get("actual_due_within_week_collections") or 0.0) for wd in active_weeks)
        actual_month_overdue = sum((wd.get("actual_overdue_collections") or 0.0) for wd in active_weeks)
        actual_month_collections = sum((wd.get("actual_collections") or 0.0) for wd in active_weeks)
        actual_month_od_utilization = sum((wd.get("actual_od_utilization") or 0.0) for wd in active_weeks)
        actual_month_total_inflows = sum((wd.get("actual_total_inflows") or 0.0) for wd in active_weeks)
        actual_month_by_category = (
            pd.DataFrame([wd["actual_by_category"] for wd in active_weeks])
            .fillna(0.0)
            .sum()
            .reindex(outflow_sheet_order, fill_value=0.0)
            if active_weeks
            else pd.Series(0.0, index=outflow_sheet_order)
        )
        actual_month_outflows = actual_month_by_category.sum()
        actual_month_net = sum((wd.get("actual_net") or 0.0) for wd in active_weeks)
        actual_month_ending = active_weeks[-1]["ending_actual"] if active_weeks else opening_balance
        actual_month_ending_od = (
            active_weeks[-1].get("ending_actual_od_outstanding")
            if active_weeks
            else max(opening_od_utilization, 0.0)
        )
        actual_month_closing = (
            actual_month_ending - (actual_month_ending_od or 0.0)
            if actual_month_ending is not None
            else None
        )

    table_rows = []
    for line_item in line_items:
        row = {"Line Item": line_item}
        for week in weeks:
            row[week["key"]] = week_values[line_item].get(week["key"])
        if line_item == "Inflows":
            row["Fcst (Selected Month)"] = None
            row["Actual (if available)"] = None
        elif line_item == "   Advance":
            row["Fcst (Selected Month)"] = monthly_plan_advance
            row["Actual (if available)"] = actual_month_advance
        elif line_item == "   Receivables":
            row["Fcst (Selected Month)"] = monthly_plan_receivables
            row["Actual (if available)"] = actual_month_receivables
        elif line_item == "      Due Within Week":
            row["Fcst (Selected Month)"] = monthly_plan_due_within_week
            row["Actual (if available)"] = actual_month_due_within_week
        elif line_item == "      Overdue":
            row["Fcst (Selected Month)"] = monthly_plan_overdue
            row["Actual (if available)"] = actual_month_overdue
        elif line_item == "Total Inflows":
            row["Fcst (Selected Month)"] = monthly_plan_total_inflows
            row["Actual (if available)"] = actual_month_total_inflows
        elif line_item == "Cash Outflows":
            row["Fcst (Selected Month)"] = None
            row["Actual (if available)"] = None
        elif line_item.startswith("   "):
            category = line_item[3:]
            row["Fcst (Selected Month)"] = monthly_plan_by_category.get(category, 0.0)
            row["Actual (if available)"] = (
                actual_month_by_category.get(category, 0.0) if actual_data else None
            )
        elif line_item == "Total Outflows":
            row["Fcst (Selected Month)"] = monthly_plan_outflows
            row["Actual (if available)"] = actual_month_outflows
        elif line_item == "Net Movement of Cash":
            row["Fcst (Selected Month)"] = monthly_plan_net
            row["Actual (if available)"] = actual_month_net
        elif line_item == "Opening Balance incl. OD Opening O/S":
            row["Fcst (Selected Month)"] = monthly_plan_opening_incl_od
            row["Actual (if available)"] = monthly_plan_opening_incl_od if actual_data else None
        elif line_item == "Closing Balance":
            row["Fcst (Selected Month)"] = monthly_plan_closing
            row["Actual (if available)"] = actual_month_closing
        table_rows.append(row)

    matrix = pd.DataFrame(table_rows)
    return {
        "weekly_table": matrix,
        "week_meta": week_meta,
        "week_columns": [week["key"] for week in weeks],
        "monthly_forecast": monthly_forecast,
        "monthly_inflow_detail": monthly_inflow_detail,
        "monthly_outflow_detail": monthly_outflow_detail,
        "plan_receivables": plan_receivables,
        "plan_outflows": plan_outflows,
        "selected_month_start": month_start,
        "selected_month_end": month_end,
        "selected_period_start": balance_date,
        "selected_month_plan": {
            "opening": opening_balance,
            "collections": monthly_plan_collections,
            "outflows": monthly_plan_outflows,
            "net": monthly_plan_net,
            "ending": monthly_plan_ending,
            "opening_od_utilization": max(opening_od_utilization, 0.0),
            "od_movement": monthly_plan_od_utilization,
            "od_utilization": monthly_plan_ending_od,
            "available_od_limit": monthly_plan_headroom,
            "ending_excl_od": monthly_plan_ending - monthly_plan_ending_od,
            "od_interest_est": monthly_plan_od_interest,
            "od_buffer_balance": od_buffer_balance,
        },
    }


# ══════════════════════════════════════════════════════════════════════════════
#  COLLECTION ENGINE & SCENARIOS
# ══════════════════════════════════════════════════════════════════════════════

def generate_scenarios(
    receivables_df: pd.DataFrame,
    opening_balance: float,
    next_30_day_outflows: float,
    anchor_date: pd.Timestamp | None = None,
    *,
    base_probs: dict[str, float] | None = None,
    worst_probs: dict[str, float] | None = None,
    tds_rate: float = 0.0,
) -> pd.DataFrame:
    if receivables_df.empty:
        return pd.DataFrame(columns=["Scenario", "Expected Collection", "Impact on Closing Cash", "Runway"])

    derived = derive_collection_probabilities(
        receivables_df, anchor_date=anchor_date,
        base_probs=base_probs, worst_probs=worst_probs, tds_rate=tds_rate,
    )
    horizon_start = build_month_start(anchor_date or TODAY)
    horizon_end = horizon_start + pd.Timedelta(days=29)
    horizon_receivables = filter_between(derived, "forecast_collection_date", horizon_start, horizon_end)
    avg_daily_outflow = next_30_day_outflows / 30 if next_30_day_outflows else 0.0

    rows = []
    for scenario, column in [("Best Case", "expected_best"), ("Base Case", "expected_base"), ("Worst Case", "expected_worst")]:
        expected = horizon_receivables[column].sum() if not horizon_receivables.empty else 0.0
        impact = opening_balance + expected - next_30_day_outflows
        runway = np.inf if avg_daily_outflow == 0 else max(impact, 0.0) / avg_daily_outflow
        rows.append({"Scenario": scenario, "Expected Collection": expected, "Impact on Closing Cash": impact, "Runway": runway})
    return pd.DataFrame(rows)


def calculate_collection_engine(
    receivables: pd.DataFrame,
    selected_month: pd.Timestamp,
    opening_balance: float,
    next_30_day_outflows: float,
    *,
    base_probs: dict[str, float] | None = None,
    worst_probs: dict[str, float] | None = None,
    tds_rate: float = 0.0,
) -> tuple[pd.DataFrame, float, float]:
    derived = derive_collection_probabilities(
        receivables, anchor_date=selected_month,
        base_probs=base_probs, worst_probs=worst_probs, tds_rate=tds_rate,
    )
    scenario_table = generate_scenarios(
        receivables, opening_balance, next_30_day_outflows, anchor_date=selected_month,
        base_probs=base_probs, worst_probs=worst_probs, tds_rate=tds_rate,
    )
    overdue_expected = derived.loc[derived["aging_bucket"].isin(["30-60", "60+"]), "expected_base"].sum() if not derived.empty else 0.0
    sixty_plus_expected = derived.loc[derived["aging_bucket"] == "60+", "expected_base"].sum() if not derived.empty else 0.0
    total_expected = derived["expected_base"].sum() if not derived.empty else 0.0
    overdue_dependency = overdue_expected / total_expected if total_expected else 0.0
    sixty_plus_dependency = sixty_plus_expected / total_expected if total_expected else 0.0
    return scenario_table, overdue_dependency, sixty_plus_dependency


# ══════════════════════════════════════════════════════════════════════════════
#  ACTUAL vs BUDGET COMPARISON
# ══════════════════════════════════════════════════════════════════════════════

def compare_actual_vs_budget(
    weekly_cashflow: dict[str, Any],
    actual_data: dict[str, Any] | None,
) -> pd.DataFrame:
    if not actual_data:
        return pd.DataFrame(columns=["Line Item", "Planned", "Actual", "Variance", "Variance %"])

    table = weekly_cashflow["weekly_table"].copy()
    rows = []
    ending_planned = np.nan
    ending_actual = np.nan
    for _, record in table.iterrows():
        line_item = str(record["Line Item"]).strip()
        if line_item in {"Inflows", "Cash Outflows"}:
            continue
        if line_item == "Closing Balance":
            ending_planned = record["Fcst (Selected Month)"]
            ending_actual = record["Actual (if available)"]
            continue
        planned = record["Fcst (Selected Month)"]
        actual = record["Actual (if available)"]
        if pd.isna(planned) and pd.isna(actual):
            continue
        planned_value = 0.0 if pd.isna(planned) else float(planned)
        actual_value = 0.0 if pd.isna(actual) else float(actual)
        variance = actual_value - planned_value
        denominator = abs(planned) if planned not in (None, 0) and not pd.isna(planned) else np.nan
        variance_pct = abs(variance) / denominator if denominator and not pd.isna(denominator) else np.nan
        rows.append({"Line Item": line_item, "Planned": planned, "Actual": actual, "Variance": variance, "Variance %": variance_pct})

    if not pd.isna(ending_planned) or not pd.isna(ending_actual):
        planned_value = 0.0 if pd.isna(ending_planned) else float(ending_planned)
        actual_value = 0.0 if pd.isna(ending_actual) else float(ending_actual)
        variance = actual_value - planned_value
        denominator = abs(ending_planned) if ending_planned not in (None, 0) and not pd.isna(ending_planned) else np.nan
        variance_pct = abs(variance) / denominator if denominator and not pd.isna(denominator) else np.nan
        rows.append({"Line Item": "Ending Cash Balance", "Planned": ending_planned, "Actual": ending_actual, "Variance": variance, "Variance %": variance_pct})
    return pd.DataFrame(rows)


def summarize_line_dates(series: pd.Series) -> str:
    dates = sorted({pd.Timestamp(v).normalize() for v in series.dropna()})
    if not dates:
        return ""
    if len(dates) == 1:
        return format_date(dates[0])
    return f"{format_date(dates[0])} to {format_date(dates[-1])}"


def summarize_inflow_detail(df: pd.DataFrame, *, value_column: str, value_label: str) -> pd.DataFrame:
    if df.empty:
        return pd.DataFrame(columns=["Inflow Sheet", "Billing Type", "Client", "Due Date", "Forecast Date", value_label, "Count of Invoices"])
    working_df = df.copy()
    working_df["_summary_line"] = build_summary_label(working_df, "counterparty", "description", "Collections")
    inflow_sheets = (
        working_df.get("source_sheet", pd.Series("Inflows", index=working_df.index))
        .fillna("Inflows")
        .astype(str)
        .str.strip()
        .replace("", "Inflows")
    )
    working_df["_source_sheet"] = inflow_sheets
    working_df["_billing_bucket"] = (
        working_df.get("billing_bucket", pd.Series("Receivables", index=working_df.index))
        .fillna("Receivables")
        .map(normalize_billing_bucket)
    )
    return (
        working_df.groupby(["_source_sheet", "_billing_bucket", "_summary_line"], dropna=False)
        .agg(**{
            value_label: (value_column, "sum"),
            "Due Date": ("due_date", summarize_line_dates),
            "Forecast Date": ("forecast_collection_date", summarize_line_dates),
            "Count of Invoices": (value_column, "size"),
        })
        .reset_index()
        .rename(columns={"_source_sheet": "Inflow Sheet", "_billing_bucket": "Billing Type", "_summary_line": "Client"})
        .sort_values(by=["Inflow Sheet", "Billing Type", value_label, "Client"], ascending=[True, True, False, True], kind="stable")
        .reset_index(drop=True)
    )


def summarize_outflow_detail(df: pd.DataFrame, *, value_label: str = "Amount") -> pd.DataFrame:
    if df.empty:
        return pd.DataFrame(columns=["Category", "Vendor", "Payment Window", value_label, "Source Rows"])
    working_df = df.copy()
    working_df["_summary_line"] = build_summary_label(working_df, "vendor_name", "description", "Outflows")
    return (
        working_df.groupby(["sheet_name", "_summary_line"], dropna=False)
        .agg(**{
            value_label: ("amount", "sum"),
            "Payment Window": ("date", summarize_line_dates),
            "Source Rows": ("amount", "size"),
        })
        .reset_index()
        .rename(columns={"sheet_name": "Category", "_summary_line": "Vendor"})
        .sort_values(by=["Category", value_label, "Vendor"], ascending=[True, False, True], kind="stable")
        .reset_index(drop=True)
    )


def prepare_inflow_line_item_table(df: pd.DataFrame) -> pd.DataFrame:
    inflows = round_numeric_columns(
        df.copy(),
        percent_columns=["base_probability", "best_probability", "worst_probability"],
    )
    for column in ["billing_type", "billing_bucket"]:
        if column in inflows.columns:
            inflows[column] = inflows[column].map(normalize_billing_bucket)
    if inflows.empty:
        return pd.DataFrame(columns=[
            "Invoice Date", "Tentative Collection Date", "Forecast Collection Date", "Due Date",
            "Billing Type", "Billing Bucket", "Client", "Reference",
            "Invoice Amount", "Aging Bucket", "Base Probability",
            "Expected Base Collection",
        ])
    keep_cols = [
        "invoice_date", "tentative_collection_date", "forecast_collection_date", "due_date",
        "source_sheet", "billing_type", "billing_bucket", "counterparty", "description", "amount", "aging_bucket",
        "base_probability", "best_probability", "worst_probability",
        "expected_base", "expected_best", "expected_worst",
    ]
    return inflows[[c for c in keep_cols if c in inflows.columns]].rename(columns={
        "invoice_date": "Invoice Date",
        "tentative_collection_date": "Tentative Collection Date",
        "forecast_collection_date": "Forecast Collection Date",
        "due_date": "Due Date",
        "source_sheet": "Inflow Sheet",
        "billing_type": "Billing Type",
        "billing_bucket": "Billing Bucket",
        "counterparty": "Client",
        "description": "Reference",
        "amount": "Invoice Amount",
        "aging_bucket": "Aging Bucket",
        "base_probability": "Base Probability",
        "best_probability": "Best Probability",
        "worst_probability": "Worst Probability",
        "expected_base": "Expected Base Collection",
        "expected_best": "Expected Best Collection",
        "expected_worst": "Expected Worst Collection",
    })


def prepare_outflow_line_item_table(df: pd.DataFrame) -> pd.DataFrame:
    outflows = round_numeric_columns(df.copy())
    if outflows.empty:
        return pd.DataFrame(columns=["Category", "Vendor", "Particular", "Payment Date", "Amount"])
    return outflows[[c for c in ["sheet_name", "vendor_name", "description", "date", "amount"] if c in outflows.columns]].rename(columns={
        "sheet_name": "Category",
        "vendor_name": "Vendor",
        "description": "Particular",
        "date": "Payment Date",
        "amount": "Amount",
    })


def resolve_weekly_drilldown_selection(
    weekly_cashflow: dict[str, Any],
    selected_month: pd.Timestamp,
) -> dict[str, Any] | None:
    params = get_query_params()
    selected_scope = params.get("weekly_drill_scope", "")
    if selected_scope != selected_month.strftime("%Y-%m"):
        return None

    week_key = params.get("weekly_drill_week", "")
    line_item = params.get("weekly_drill_line", "")
    week = next((entry for entry in weekly_cashflow["week_meta"] if entry["key"] == week_key), None)
    if week is None:
        return None

    outflow_categories = list(week["plan_by_category"].index)
    drill_config = classify_weekly_drilldown_line_item(line_item, outflow_categories)
    if drill_config is None:
        return None

    if drill_config["section"] == "inflows":
        inflow_detail = week["plan_receipts_detail"].copy()
        if drill_config["mode"] in {"advance", "receivables", "due_within_week", "overdue"}:
            bucket_masks = classify_weekly_receivable_masks(
                inflow_detail,
                week["effective_start"],
                week["end"],
            )
            inflow_detail = inflow_detail.loc[bucket_masks[drill_config["mode"]]].copy()
        return {
            "title": f"{week['key']} - {drill_config['label']}",
            "subtitle": week["label"],
            "summary_df": summarize_inflow_detail(
                inflow_detail,
                value_column="expected_base",
                value_label="Expected Collection",
            ),
            "summary_currency_columns": ["Expected Collection"],
            "detail_df": prepare_inflow_line_item_table(inflow_detail),
            "detail_currency_columns": ["Invoice Amount", "Expected Base Collection", "Expected Best Collection", "Expected Worst Collection"],
            "detail_percent_columns": ["Base Probability", "Best Probability", "Worst Probability"],
            "empty_message": "No inflow line items found for this week selection.",
        }

    outflow_detail = week["plan_outflows_detail"].copy()
    if drill_config["mode"] == "category":
        outflow_detail = outflow_detail.loc[
            outflow_detail["sheet_name"].astype(str).map(standardize_label) == standardize_label(drill_config["category"])
        ].copy()
    return {
        "title": f"{week['key']} - {drill_config['label']}",
        "subtitle": week["label"],
        "summary_df": summarize_outflow_detail(outflow_detail),
        "summary_currency_columns": ["Amount"],
        "detail_df": prepare_outflow_line_item_table(outflow_detail),
        "detail_currency_columns": ["Amount"],
        "detail_percent_columns": [],
        "empty_message": "No outflow line items found for this week selection.",
    }


def render_weekly_drilldown_selection(
    weekly_cashflow: dict[str, Any],
    selected_month: pd.Timestamp,
) -> None:
    st.markdown("**Weekly Cell Drill-Down**")
    selection = resolve_weekly_drilldown_selection(weekly_cashflow, selected_month)
    if selection is None:
        st.caption("Click a weekly amount in Advance, Receivables, Due Within Week, Overdue, Total Inflows, Total Outflows, or an outflow category row to inspect supporting line items.")
        return

    title_col, action_col = st.columns([5, 1])
    with title_col:
        st.caption(selection["subtitle"])
        st.markdown(f"**{selection['title']}**")
    with action_col:
        if st.button("Clear", key="clear_weekly_drilldown", use_container_width=True):
            clear_weekly_drilldown_query_params()
            st.rerun()

    if selection["detail_df"].empty:
        st.info(selection["empty_message"])
        return

    st.caption("Grouped summary")
    render_simple_table(
        selection["summary_df"],
        currency_columns=selection["summary_currency_columns"],
        enable_sort=True,
        sort_key_prefix="weekly_drilldown_summary",
        sortable_columns=selection["summary_df"].columns.tolist(),
    )
    with st.expander("Show underlying line items", expanded=False):
        render_simple_table(
            selection["detail_df"],
            currency_columns=selection["detail_currency_columns"],
            percent_columns=selection["detail_percent_columns"],
            enable_sort=True,
            sort_key_prefix="weekly_drilldown_detail",
            sortable_columns=selection["detail_df"].columns.tolist(),
        )


def build_line_level_variance(
    weekly_cashflow: dict[str, Any],
    actual_data: dict[str, Any] | None,
    *,
    base_probs: dict[str, float] | None = None,
    worst_probs: dict[str, float] | None = None,
    tds_rate: float = 0.0,
) -> pd.DataFrame:
    columns = [
        "Section", "Category", "Line Item",
        "Planned Date", "Actual Date", "Status",
        "Planned", "Actual", "Variance", "Variance %",
    ]
    if not actual_data:
        return pd.DataFrame(columns=columns)

    month_start = weekly_cashflow.get("selected_period_start", weekly_cashflow["selected_month_start"])
    month_end = weekly_cashflow["selected_month_end"]

    plan_receivables = filter_between(weekly_cashflow["plan_receivables"], "forecast_collection_date", month_start, month_end)
    actual_receivables_view = build_actual_receivables_view(
        actual_data, anchor_date=month_start,
        base_probs=base_probs, worst_probs=worst_probs, tds_rate=tds_rate,
    )
    actual_receivables = filter_between(actual_receivables_view, "comparison_date", month_start, month_end)
    plan_outflows = filter_between(weekly_cashflow["plan_outflows"], "date", month_start, month_end)
    actual_outflows = filter_between(actual_data["outflows"], "date", month_start, month_end)
    inflow_category_column = "source_sheet" if "source_sheet" in plan_receivables.columns or "source_sheet" in actual_receivables.columns else None

    def aggregate_lines(
        df, *, section, category_column, summary_column, summary_fallback_column,
        date_column, value_column, value_label, date_label, category_label=None,
    ):
        if df.empty:
            return pd.DataFrame(columns=["Section", "Category", "Line Item", date_label, value_label])
        working_df = df.copy()
        working_df["_summary_line"] = build_summary_label(
            working_df, summary_column, summary_fallback_column, category_label or section
        )
        if category_column is None:
            working_df["_category_label"] = category_label or section
            grp_cols = ["_category_label", "_summary_line"]
            rename_src = "_category_label"
        else:
            grp_cols = [category_column, "_summary_line"]
            rename_src = category_column
        aggregated = (
            working_df.groupby(grp_cols, dropna=False)
            .agg(**{value_label: (value_column, "sum"), date_label: (date_column, summarize_line_dates)})
            .reset_index()
            .rename(columns={rename_src: "Category", "_summary_line": "Line Item"})
        )
        aggregated.insert(0, "Section", section)
        aggregated["Category"] = aggregated["Category"].fillna(section)
        aggregated["Line Item"] = aggregated["Line Item"].fillna("").astype(str).str.strip()
        return aggregated

    planned_lines = pd.concat([
        aggregate_lines(plan_receivables, section="Collections", category_column=inflow_category_column,
                        summary_column="counterparty", summary_fallback_column="description",
                        date_column="forecast_collection_date", value_column="expected_base",
                        value_label="Planned", date_label="Planned Date", category_label="Inflows"),
        aggregate_lines(plan_outflows, section="Outflows", category_column="sheet_name",
                        summary_column="vendor_name", summary_fallback_column="description",
                        date_column="date", value_column="amount",
                        value_label="Planned", date_label="Planned Date"),
    ], ignore_index=True)

    actual_lines = pd.concat([
        aggregate_lines(actual_receivables, section="Collections", category_column=inflow_category_column,
                        summary_column="counterparty", summary_fallback_column="description",
                        date_column="comparison_date", value_column="comparison_amount",
                        value_label="Actual", date_label="Actual Date", category_label="Inflows"),
        aggregate_lines(actual_outflows, section="Outflows", category_column="sheet_name",
                        summary_column="vendor_name", summary_fallback_column="description",
                        date_column="date", value_column="amount",
                        value_label="Actual", date_label="Actual Date"),
    ], ignore_index=True)

    merged = planned_lines.merge(actual_lines, on=["Section", "Category", "Line Item"], how="outer")
    if merged.empty:
        return pd.DataFrame(columns=columns)

    merged["Planned"] = merged["Planned"].fillna(0.0)
    merged["Actual"] = merged["Actual"].fillna(0.0)
    merged["Planned Date"] = merged["Planned Date"].fillna("")
    merged["Actual Date"] = merged["Actual Date"].fillna("")
    merged["Variance"] = merged["Actual"] - merged["Planned"]
    merged["Variance %"] = np.where(
        merged["Planned"].abs() > 0,
        merged["Variance"].abs() / merged["Planned"].abs(),
        np.nan,
    )
    merged["Status"] = np.select(
        [
            (merged["Planned"] > 0) & (merged["Actual"] > 0),
            (merged["Planned"] > 0) & (merged["Actual"] <= 0),
            (merged["Planned"] <= 0) & (merged["Actual"] > 0),
        ],
        ["Matched", "Planned only", "Actual only"],
        default="Review",
    )
    merged["_abs_variance"] = merged["Variance"].abs()
    merged = merged.sort_values(
        by=["Section", "Category", "_abs_variance", "Line Item"],
        ascending=[True, True, False, True],
    ).drop(columns="_abs_variance")
    return merged[columns].reset_index(drop=True)


# ══════════════════════════════════════════════════════════════════════════════
#  FORMATTERS
# ══════════════════════════════════════════════════════════════════════════════

def format_currency(value: Any) -> str:
    if value is None or pd.isna(value):
        return ""
    return f"{value:,.0f}"


def format_percent(value: Any) -> str:
    if value is None or pd.isna(value):
        return ""
    return f"{value:.0%}"


def format_runway(value: Any) -> str:
    if value is None or pd.isna(value):
        return ""
    if value == np.inf:
        return "No burn"
    return f"{value:,.0f} days"


def negative_red(value: Any) -> str:
    if value is None or pd.isna(value):
        return ""
    return "color: #b91c1c;" if float(value) < 0 else ""


def round_numeric_columns(df: pd.DataFrame, percent_columns: list[str] | None = None) -> pd.DataFrame:
    rounded = df.copy()
    percent_set = set(percent_columns or [])
    for column in rounded.columns:
        if column in percent_set or not pd.api.types.is_numeric_dtype(rounded[column]):
            continue
        rounded[column] = rounded[column].apply(
            lambda v: int(round(float(v))) if pd.notna(v) else v
        )
    return rounded


def apply_excel_number_formats(worksheet) -> None:
    percent_headers = {"Variance %"}
    for column_cells in worksheet.columns:
        header = str(column_cells[0].value or "")
        if not header:
            continue
        is_percent_column = ("Probability" in header) or (header in percent_headers)
        for cell in column_cells[1:]:
            if isinstance(cell.value, (datetime, date, pd.Timestamp)):
                cell.number_format = "DD-MM-YYYY"
                continue
            if cell.value is None or isinstance(cell.value, bool) or not isinstance(cell.value, Number):
                continue
            cell.number_format = "0%" if is_percent_column else "#,##0"


def bold_totals(row: pd.Series) -> list[str]:
    line_item = str(row.get("Line Item", "")).strip()
    return [
        "font-weight: 700;"
        if line_item in TOTAL_ROWS
        else ""
        for _ in row
    ]


# ══════════════════════════════════════════════════════════════════════════════
#  SORT CONTROLS
# ══════════════════════════════════════════════════════════════════════════════

def apply_sort_controls(
    df: pd.DataFrame,
    *,
    key_prefix: str,
    sortable_columns: list[str] | None = None,
    default_sort_by: str | None = None,
    default_descending: bool = False,
    preserve_index: bool = False,
) -> pd.DataFrame:
    if df.empty or len(df) <= 1:
        return df
    available_columns = [c for c in (sortable_columns or df.columns.tolist()) if c in df.columns]
    if not available_columns:
        return df
    default_label = "Default order"
    sort_options = [default_label] + available_columns
    default_index = sort_options.index(default_sort_by) if default_sort_by in available_columns else 0
    c1, c2 = st.columns([1.35, 1])
    selected_column = c1.selectbox("Sort by", sort_options, index=default_index, key=f"{key_prefix}_sort_by")
    selected_order = c2.selectbox(
        "Order", ["Ascending", "Descending"],
        index=1 if default_descending else 0,
        key=f"{key_prefix}_sort_order",
        disabled=selected_column == default_label,
    )
    if selected_column == default_label:
        return df.copy() if preserve_index else df.reset_index(drop=True)
    sorted_df = df.sort_values(
        by=selected_column, ascending=(selected_order == "Ascending"), na_position="last", kind="stable"
    )
    return sorted_df if preserve_index else sorted_df.reset_index(drop=True)


def sort_financial_matrix(
    df: pd.DataFrame,
    *,
    key_prefix: str,
    sortable_columns: list[str] | None = None,
    preserve_total_rows: bool = True,
) -> pd.DataFrame:
    if df.empty or len(df) <= 1:
        return df
    available_columns = [c for c in (sortable_columns or df.columns.tolist()) if c in df.columns]
    if not available_columns:
        return df
    default_label = "Default order"
    sort_options = [default_label] + available_columns
    c1, c2 = st.columns([1.35, 1])
    selected_column = c1.selectbox("Sort by", sort_options, index=0, key=f"{key_prefix}_sort_by")
    selected_order = c2.selectbox(
        "Order", ["Ascending", "Descending"], index=0, key=f"{key_prefix}_sort_order",
        disabled=selected_column == default_label,
    )
    if selected_column == default_label:
        return df.reset_index(drop=True)
    totals_mask = (
        df["Line Item"].astype(str).str.strip().isin(TOTAL_ROWS)
        if preserve_total_rows and "Line Item" in df.columns
        else pd.Series(False, index=df.index)
    )
    pinned = df.loc[totals_mask].copy()
    sortable = df.loc[~totals_mask].copy()
    if sortable.empty:
        return df.reset_index(drop=True)
    sorted_rows = sortable.sort_values(
        by=selected_column, ascending=(selected_order == "Ascending"), na_position="last", kind="stable"
    )
    return (sorted_rows if pinned.empty else pd.concat([pinned, sorted_rows], ignore_index=True))


# ══════════════════════════════════════════════════════════════════════════════
#  TABLE RENDERERS
# ══════════════════════════════════════════════════════════════════════════════

def render_financial_table(
    df: pd.DataFrame,
    numeric_columns: list[str],
    height: int | None = None,
    *,
    enable_sort: bool = False,
    sort_key_prefix: str | None = None,
    sortable_columns: list[str] | None = None,
    preserve_total_rows: bool = True,
) -> None:
    if df.empty:
        st.info("No records available for this view.")
        return
    display_df = df.copy()
    if enable_sort and sort_key_prefix:
        display_df = sort_financial_matrix(
            display_df, key_prefix=sort_key_prefix,
            sortable_columns=sortable_columns, preserve_total_rows=preserve_total_rows,
        )
    display_df = format_date_columns(display_df)
    styler = (
        display_df.style.hide(axis="index")
        .format({col: format_currency for col in numeric_columns}, na_rep="")
        .apply(bold_totals, axis=1)
        .applymap(negative_red, subset=numeric_columns)
        .set_properties(subset=["Line Item"], **{"text-align": "left", "white-space": "pre"})
        .set_properties(subset=numeric_columns, **{"text-align": "right"})
        .set_table_styles([
            {"selector": "table", "props": [("width", "100%"), ("border-collapse", "collapse"), ("font-size", "0.95rem")]},
            {"selector": "th", "props": [("text-align", "right"), ("padding", "8px 10px"), ("border-bottom", "1px solid #d4d4d8")]},
            {"selector": "th.col_heading.level0.col0", "props": [("text-align", "left")]},
            {"selector": "td", "props": [("padding", "8px 10px"), ("border-bottom", "1px solid #ececf1")]},
        ])
    )
    st.markdown(styler.to_html(), unsafe_allow_html=True)


def render_weekly_financial_table(
    weekly_cashflow: dict[str, Any],
    selected_month: pd.Timestamp,
    balance_as_of_date: pd.Timestamp,
) -> None:
    df = weekly_cashflow["weekly_table"].copy()
    if df.empty:
        st.info("No records available for this view.")
        return

    week_columns = set(weekly_cashflow["week_columns"])
    week_header_map = {
        week["key"]: f'{week["key"]}<br><span style="font-weight:400; font-size:0.82rem;">'
        f'{compact_day_date_label(week["start"])} - {compact_day_date_label(week["end"])}</span>'
        for week in weekly_cashflow["week_meta"]
    }
    outflow_categories = weekly_cashflow["plan_outflows"]["sheet_name"].dropna().astype(str).unique().tolist()
    scope = selected_month.strftime("%Y-%m")
    display_df = format_date_columns(df)
    header_columns = display_df.columns.tolist()

    html_rows: list[str] = [
        '<div style="width:100%; overflow-x:auto;">',
        '<table style="width:100%; border-collapse:collapse; font-size:0.95rem;">',
        "<thead><tr>",
    ]
    for idx, column in enumerate(header_columns):
        align = "left" if idx == 0 else "right"
        header_label = week_header_map.get(column, html.escape(str(column)))
        html_rows.append(
            f'<th style="text-align:{align}; padding:8px 10px; border-bottom:1px solid #d4d4d8; white-space:nowrap;">'
            f"{header_label}</th>"
        )
    html_rows.append("</tr></thead><tbody>")

    for _, row in display_df.iterrows():
        line_item = str(row["Line Item"])
        drill_config = classify_weekly_drilldown_line_item(line_item, outflow_categories)
        row_label = line_item.strip()
        row_weight = "600" if row_label in TOTAL_ROWS else "400"
        html_rows.append("<tr>")
        for column in header_columns:
            if column == "Line Item":
                html_rows.append(
                    '<td style="padding:8px 10px; border-bottom:1px solid #ececf1; text-align:left; '
                    f'white-space:pre; font-weight:{row_weight};">{html.escape(line_item)}</td>'
                )
                continue

            raw_value = df.loc[row.name, column]
            formatted_value = format_currency(raw_value) if pd.notna(raw_value) else ""
            cell_color = "#b91c1c" if pd.notna(raw_value) and float(raw_value) < 0 else "inherit"
            cell_value_html = html.escape(formatted_value)

            if column in week_columns and formatted_value and drill_config is not None:
                drill_href = build_weekly_drilldown_href(
                    scope,
                    column,
                    line_item,
                    selected_month,
                    balance_as_of_date,
                )
                cell_value_html = (
                    f'<a href="{html.escape(drill_href)}" target="_self" '
                    'style="color:inherit; text-decoration:none; font-weight:inherit;">'
                    f"{html.escape(formatted_value)}</a>"
                )

            html_rows.append(
                '<td style="padding:8px 10px; border-bottom:1px solid #ececf1; text-align:right; '
                f'color:{cell_color}; font-weight:{row_weight};">{cell_value_html}</td>'
            )
        html_rows.append("</tr>")

    html_rows.append("</tbody></table></div>")
    st.markdown("".join(html_rows), unsafe_allow_html=True)
    st.caption("Click a weekly amount cell to view the supporting line items below.")


def render_simple_table(
    df: pd.DataFrame,
    currency_columns: list[str] | None = None,
    percent_columns: list[str] | None = None,
    *,
    enable_sort: bool = False,
    sort_key_prefix: str | None = None,
    sortable_columns: list[str] | None = None,
    default_sort_by: str | None = None,
    default_descending: bool = False,
) -> None:
    if df.empty:
        st.info("No records available.")
        return
    currency_columns = currency_columns or []
    percent_columns = percent_columns or []
    display_df = df.copy()
    if enable_sort and sort_key_prefix:
        display_df = apply_sort_controls(
            display_df, key_prefix=sort_key_prefix, sortable_columns=sortable_columns,
            default_sort_by=default_sort_by, default_descending=default_descending,
        )
    display_df = format_date_columns(display_df)
    styler = display_df.style.hide(axis="index")
    formatters: dict[str, Any] = {}
    if currency_columns:
        formatters.update({col: format_currency for col in currency_columns})
        styler = styler.applymap(negative_red, subset=currency_columns)
    if percent_columns:
        formatters.update({col: format_percent for col in percent_columns})
    if formatters:
        styler = styler.format(formatters, na_rep="")
    styler = styler.set_table_styles([
        {"selector": "table", "props": [("width", "max-content"), ("min-width", "100%"), ("border-collapse", "collapse"), ("font-size", "0.94rem")]},
        {"selector": "th", "props": [("text-align", "left"), ("padding", "8px 10px"), ("border-bottom", "1px solid #d4d4d8"), ("vertical-align", "top"), ("white-space", "nowrap")]},
        {"selector": "td", "props": [("padding", "8px 10px"), ("border-bottom", "1px solid #ececf1"), ("vertical-align", "top"), ("white-space", "normal"), ("word-break", "break-word")]},
    ])
    st.markdown(
        f'<div style="width:100%; overflow-x:auto;">{styler.to_html()}</div>',
        unsafe_allow_html=True,
    )


# ══════════════════════════════════════════════════════════════════════════════
#  CFO ALERT BANNER  (new)
# ══════════════════════════════════════════════════════════════════════════════

def render_cash_alert_banner(
    projected_ending: float,
    min_cash_threshold: float,
    od_utilization: float,
    od_sanctioned_limit: float,
    od_interest_est: float,
) -> None:
    """
    Display a red/amber alert strip when cash < threshold or OD is at limit.
    Only visible when there is an active risk condition.
    """
    alerts: list[tuple[str, str]] = []   # (severity, message)

    if min_cash_threshold > 0 and projected_ending < min_cash_threshold:
        if projected_ending < 0:
            alerts.append(("red", f"CRITICAL: Projected month-end cash is negative ({format_currency(projected_ending)}). Immediate action required."))
        else:
            gap = min_cash_threshold - projected_ending
            alerts.append(("amber", f"WARNING: Projected closing cash ({format_currency(projected_ending)}) is below the minimum floor of {format_currency(min_cash_threshold)} — shortfall of {format_currency(gap)}."))

    if od_sanctioned_limit > 0 and od_utilization >= 0.90 * od_sanctioned_limit:
        util_pct = od_utilization / od_sanctioned_limit * 100
        alerts.append(("amber", f"OD HEADROOM TIGHT: Utilising {util_pct:.0f}% of sanctioned limit ({format_currency(od_utilization)} of {format_currency(od_sanctioned_limit)})."))

    if od_interest_est > 0:
        alerts.append(("info", f"Estimated OD interest for this month: {format_currency(od_interest_est)} based on average OD outstanding at the configured rate."))

    for severity, message in alerts:
        colour_map = {
            "red": ("#fef2f2", "#991b1b", "#fecaca"),
            "amber": ("#fffbeb", "#92400e", "#fde68a"),
            "info": ("#eff6ff", "#1e40af", "#bfdbfe"),
        }
        bg, text, border = colour_map.get(severity, colour_map["info"])
        st.markdown(
            f"""
            <div style="border:1px solid {border}; border-left:4px solid {text};
                        background:{bg}; color:{text}; padding:12px 16px;
                        margin:8px 0; border-radius:4px; font-size:0.94rem; font-weight:600;">
                {message}
            </div>
            """,
            unsafe_allow_html=True,
        )


# ══════════════════════════════════════════════════════════════════════════════
#  KPI SECTION
# ══════════════════════════════════════════════════════════════════════════════

def render_kpis(
    opening_balance: float,
    weekly_cashflow: dict[str, Any],
    scenario_table: pd.DataFrame,
    next_30_day_requirement: float,
    overdue_dependency: float,
    sixty_plus_dependency: float,
    *,
    min_cash_threshold: float = 0.0,
    od_sanctioned_limit: float = 0.0,
) -> None:
    plan = weekly_cashflow["selected_month_plan"]
    base_case = scenario_table.loc[scenario_table["Scenario"] == "Base Case"].iloc[0] if not scenario_table.empty else None
    runway_days = base_case["Runway"] if base_case is not None else np.inf

    # ── Alert banners ────────────────────────────────────────────────────────
    render_cash_alert_banner(
        projected_ending=plan["ending"],
        min_cash_threshold=min_cash_threshold,
        od_utilization=plan.get("od_utilization", 0.0),
        od_sanctioned_limit=od_sanctioned_limit,
        od_interest_est=plan.get("od_interest_est", 0.0),
    )

    # ── Primary KPI row ──────────────────────────────────────────────────────
    cols = st.columns(5)
    cols[0].metric("Opening Cash", format_currency(opening_balance))
    cols[1].metric("Total Inflows (Month)", format_currency(plan["collections"]))
    cols[2].metric("Total Outflows (Month)", format_currency(plan["outflows"]))
    cols[3].metric("Month-End Cash (Planned)", format_currency(plan["ending"]))
    cols[4].metric("Runway", format_runway(runway_days))

    # ── OD / credit-line KPI row ─────────────────────────────────────────────
    od_util = plan.get("od_utilization", 0.0) or 0.0
    opening_od = plan.get("opening_od_utilization", 0.0) or 0.0
    if od_sanctioned_limit > 0 or od_util > 0 or opening_od > 0:
        od_headroom = plan.get("available_od_limit", max(od_sanctioned_limit - od_util, 0.0)) or 0.0
        od_interest = plan.get("od_interest_est", 0.0) or 0.0
        od_cols = st.columns(5)
        od_cols[0].metric("OD Sanctioned Limit", format_currency(od_sanctioned_limit) if od_sanctioned_limit > 0 else "—")
        od_cols[1].metric("Opening OD Outstanding", format_currency(opening_od))
        od_cols[2].metric("Closing OD Outstanding", format_currency(od_util))
        od_cols[3].metric("Available OD Headroom", format_currency(od_headroom))
        od_cols[4].metric("Est. OD Interest Cost", format_currency(od_interest) if od_interest > 0 else "Nil")

    # ── Decision box ─────────────────────────────────────────────────────────
    statement, actions = build_decision_box(
        runway_days, overdue_dependency, sixty_plus_dependency, plan["ending"],
        min_cash_threshold=min_cash_threshold,
    )
    bullets = "".join(f"<li>{action}</li>" for action in actions[:3])
    st.markdown(
        f"""
        <div style="border:1px solid #d4d4d8; border-left:4px solid #111827;
                    padding:16px 18px; margin-top:10px; margin-bottom:18px; background:#ffffff;">
            <div style="font-size:1.0rem; font-weight:700; margin-bottom:8px;">Decision Box</div>
            <div style="font-size:0.98rem; margin-bottom:10px;">{statement}</div>
            <ul style="margin:0; padding-left:18px;">{bullets}</ul>
        </div>
        """,
        unsafe_allow_html=True,
    )


def build_decision_box(
    runway_days: float,
    overdue_dependency: float,
    sixty_plus_dependency: float,
    projected_closing_cash: float,
    *,
    min_cash_threshold: float = 0.0,
) -> tuple[str, list[str]]:
    if runway_days != np.inf and runway_days <= 30:
        sentence = f"Cash will run out in approximately {max(runway_days, 0):.0f} days. Primary risk is near-term liquidity compression."
    elif projected_closing_cash < 0:
        sentence = "Projected month-end cash is negative. Primary risk is a forecast shortfall against committed payments."
    elif min_cash_threshold > 0 and projected_closing_cash < min_cash_threshold:
        sentence = f"Projected closing cash ({format_currency(projected_closing_cash)}) is below the minimum operating floor ({format_currency(min_cash_threshold)}). Buffer must be rebuilt before discretionary payments."
    elif overdue_dependency > 0.30:
        sentence = "Liquidity is dependent on overdue collections. Primary risk is slippage in already aged receivables."
    else:
        sentence = "Liquidity is currently manageable. Primary watchpoint is preserving forecast collections against committed outflows."

    actions: list[str] = []
    if overdue_dependency > 0.30:
        actions.append("Escalate overdue collections immediately and track owner-level recovery actions.")
    if sixty_plus_dependency > 0.30:
        actions.append("Stress-test the plan assuming 60+ day receivables slip beyond the month.")
    if projected_closing_cash < 0 or (runway_days != np.inf and runway_days <= 30):
        actions.append("Sequence discretionary payments after payroll, statutory dues, and critical vendors.")
    if min_cash_threshold > 0 and projected_closing_cash < min_cash_threshold:
        actions.append(f"Preserve minimum cash floor of {format_currency(min_cash_threshold)}. Review deferred or discretionary payments first.")
    if not actions:
        actions.append("Monitor weekly receipts against the base case and refresh the forecast at each weekly close.")
    return sentence, actions[:3]


# ══════════════════════════════════════════════════════════════════════════════
#  CFO COMMENTARY BOX  (new)
# ══════════════════════════════════════════════════════════════════════════════

def render_cfo_commentary(section_key: str, label: str = "CFO Commentary") -> None:
    """
    Persistent, editable commentary box. Content is stored in session state
    and included in Excel exports.
    """
    st.markdown(f"**{label}**")
    st.text_area(
        label,
        key=f"commentary_{section_key}",
        placeholder="Add narrative, key decisions, or variance explanations here. This will be included in the Excel export.",
        height=90,
        label_visibility="collapsed",
    )


# ══════════════════════════════════════════════════════════════════════════════
#  SECTION RENDERERS
# ══════════════════════════════════════════════════════════════════════════════

def render_collection_engine(
    scenario_table: pd.DataFrame,
    overdue_dependency: float,
    sixty_plus_dependency: float,
) -> None:
    scenario_display = scenario_table.copy()
    if not scenario_display.empty:
        scenario_display["Runway"] = scenario_display["Runway"].map(format_runway)
    render_simple_table(scenario_display, currency_columns=["Expected Collection", "Impact on Closing Cash"])

    flag_text = (
        "Flag: More than 30% of forecast collections are coming from 60+ day receivables."
        if sixty_plus_dependency > 0.30
        else "Flag: 60+ day receivable dependency remains within the 30% guardrail."
    )
    st.markdown(
        f"""
        <div style="border:1px solid #d4d4d8; padding:14px 16px; margin-top:12px; background:#ffffff;">
            <div><strong>% dependency on overdue collections:</strong> {format_percent(overdue_dependency)}</div>
            <div style="margin-top:6px; color:{'#b91c1c' if sixty_plus_dependency > 0.30 else '#111827'};">{flag_text}</div>
        </div>
        """,
        unsafe_allow_html=True,
    )


def render_six_month_view(
    monthly_forecast: pd.DataFrame,
    monthly_inflow_detail: pd.DataFrame,
    monthly_outflow_detail: pd.DataFrame,
    *,
    min_cash_threshold: float = 0.0,
) -> None:
    if monthly_forecast.empty:
        st.info("No monthly forecast available.")
        return

    display = monthly_forecast[["Month", "Opening Cash", "Inflows", "Outflows", "Closing Cash"]].copy()

    def style_closing_row(row: pd.Series) -> list[str]:
        closing = row["Closing Cash"]
        base = "font-weight: 700; "
        if closing < 0:
            return [base + "color: #b91c1c;" for _ in row]
        if min_cash_threshold > 0 and closing < min_cash_threshold:
            return [base + "color: #92400e;" for _ in row]
        return ["" for _ in row]

    styler = (
        display.style.hide(axis="index")
        .format({
            "Opening Cash": format_currency,
            "Inflows": format_currency,
            "Outflows": format_currency,
            "Closing Cash": format_currency,
        }, na_rep="")
        .applymap(negative_red, subset=["Closing Cash"])
        .apply(style_closing_row, axis=1)
        .set_table_styles([
            {"selector": "table", "props": [("width", "100%"), ("border-collapse", "collapse"), ("font-size", "0.95rem")]},
            {"selector": "th", "props": [("text-align", "left"), ("padding", "8px 10px"), ("border-bottom", "1px solid #d4d4d8")]},
            {"selector": "td", "props": [("padding", "8px 10px"), ("border-bottom", "1px solid #ececf1")]},
        ])
    )
    st.markdown(styler.to_html(), unsafe_allow_html=True)

    if min_cash_threshold > 0:
        breach_months = monthly_forecast.loc[monthly_forecast["Closing Cash"] < min_cash_threshold, "Month"].tolist()
        if breach_months:
            st.markdown(
                f"""
                <div style="border:1px solid #fde68a; border-left:4px solid #92400e;
                            background:#fffbeb; color:#92400e; padding:10px 14px;
                            margin:8px 0; border-radius:4px; font-size:0.9rem;">
                    <strong>Cash floor breach:</strong> {', '.join(breach_months)} — closing cash below minimum threshold of {format_currency(min_cash_threshold)}.
                </div>
                """,
                unsafe_allow_html=True,
            )

    st.markdown("**Monthly Inflow Detail**")
    render_financial_table(
        monthly_inflow_detail,
        numeric_columns=[c for c in monthly_inflow_detail.columns if c != "Line Item"],
        enable_sort=True,
        sort_key_prefix="six_month_inflow_detail",
        sortable_columns=monthly_inflow_detail.columns.tolist(),
    )
    st.markdown("**Monthly Expense Category Detail**")
    render_financial_table(
        monthly_outflow_detail,
        numeric_columns=[c for c in monthly_outflow_detail.columns if c != "Line Item"],
        enable_sort=True,
        sort_key_prefix="six_month_outflow_detail",
        sortable_columns=monthly_outflow_detail.columns.tolist(),
    )


def render_actual_vs_budget(
    plan_results: dict[str, Any],
    actual_data: dict[str, Any] | None,
    actual_file_name: str | None,
    *,
    base_probs: dict[str, float] | None = None,
    worst_probs: dict[str, float] | None = None,
    tds_rate: float = 0.0,
) -> None:
    comparison = compare_actual_vs_budget(plan_results, actual_data)
    line_variance = build_line_level_variance(
        plan_results, actual_data,
        base_probs=base_probs, worst_probs=worst_probs, tds_rate=tds_rate,
    )
    left_col, right_col = st.columns([1, 2])

    with left_col:
        st.markdown("**Actual Data**")
        if actual_file_name:
            st.success(f"Actual workbook loaded: {actual_file_name}")
            st.caption("Use the sidebar uploader to replace or clear the actual workbook.")
        else:
            st.info("Upload an actuals workbook from the sidebar to unlock variance tracking.")

    with right_col:
        st.markdown("**Comparison**")
        if comparison.empty:
            st.info("Actual data not available for comparison.")
        else:
            comparison = apply_sort_controls(
                comparison, key_prefix="actual_vs_budget_comparison",
                sortable_columns=["Line Item", "Planned", "Actual", "Variance", "Variance %"],
            )
            major_deviation_mask = comparison["Variance %"].fillna(0) > 0.10
            cash_impact = comparison.loc[comparison["Line Item"] == "Ending Cash Balance", "Variance"].sum()
            styler = (
                comparison.style.hide(axis="index")
                .format({"Planned": format_currency, "Actual": format_currency, "Variance": format_currency, "Variance %": format_percent}, na_rep="")
                .applymap(negative_red, subset=["Variance"])
                .apply(
                    lambda row: [
                        "background-color: #fef2f2;"
                        if (row["Variance %"] if not pd.isna(row["Variance %"]) else 0) > 0.10
                        else ""
                        for _ in row
                    ],
                    axis=1,
                )
                .set_table_styles([
                    {"selector": "table", "props": [("width", "100%"), ("border-collapse", "collapse"), ("font-size", "0.95rem")]},
                    {"selector": "th", "props": [("text-align", "left"), ("padding", "8px 10px"), ("border-bottom", "1px solid #d4d4d8")]},
                    {"selector": "td", "props": [("padding", "8px 10px"), ("border-bottom", "1px solid #ececf1")]},
                ])
            )
            st.markdown(styler.to_html(), unsafe_allow_html=True)
            st.caption(f"Cash impact of variance: {format_currency(cash_impact)}")

            major_deviations = comparison.loc[major_deviation_mask, ["Line Item", "Variance", "Variance %"]]
            if not major_deviations.empty:
                st.caption("Major deviations (>10%)")
                render_simple_table(
                    major_deviations,
                    currency_columns=["Variance"],
                    percent_columns=["Variance %"],
                    enable_sort=True,
                    sort_key_prefix="actual_vs_budget_major_deviations",
                    sortable_columns=["Line Item", "Variance", "Variance %"],
                )

    # ── Variance Commentary ────────────────────────────────────────────────────
    if not comparison.empty:
        render_cfo_commentary("avb_variance", label="Variance Commentary")

    if comparison.empty:
        return

    st.markdown("**Summary-Level Variance**")
    if line_variance.empty:
        st.info("No selected-month summary rows available for comparison.")
        return

    f1, f2, f3 = st.columns([1, 1, 1.4])
    section_options = ["All"] + sorted(line_variance["Section"].dropna().astype(str).unique().tolist())
    selected_section = f1.selectbox("Section", section_options, key="avb_line_section")
    cat_src = line_variance if selected_section == "All" else line_variance.loc[line_variance["Section"] == selected_section]
    cat_options = ["All"] + sorted(cat_src["Category"].dropna().astype(str).unique().tolist())
    selected_category = f2.selectbox("Category", cat_options, key="avb_line_category")
    li_src = cat_src if selected_category == "All" else cat_src.loc[cat_src["Category"] == selected_category]
    selected_line_items = f3.multiselect(
        "Summary Lines",
        options=sorted(li_src["Line Item"].dropna().astype(str).unique().tolist()),
        placeholder="All vendors or clients",
        key="avb_line_items",
    )

    c1, c2, c3, c4 = st.columns([1, 1, 0.9, 1.2])
    status_options = ["All"] + sorted(line_variance["Status"].dropna().astype(str).unique().tolist())
    selected_status = c1.selectbox("Status", status_options, key="avb_status")
    sort_order = c2.selectbox(
        "Sort By",
        ["Largest Variance", "Largest Variance %", "Largest Planned", "Largest Actual", "A-Z"],
        key="avb_sort_order",
    )
    only_exceptions = c3.checkbox("Only Exceptions", value=False, key="avb_only_exceptions")
    search_text = c4.text_input("Search", value="", placeholder="Vendor, client, or category", key="avb_search").strip().lower()

    flv = line_variance.copy()
    if selected_section != "All":
        flv = flv.loc[flv["Section"] == selected_section]
    if selected_category != "All":
        flv = flv.loc[flv["Category"] == selected_category]
    if selected_line_items:
        flv = flv.loc[flv["Line Item"].isin(selected_line_items)]
    if selected_status != "All":
        flv = flv.loc[flv["Status"] == selected_status]
    if only_exceptions:
        flv = flv.loc[(flv["Status"] != "Matched") | (flv["Variance %"].fillna(0) > 0.10)]
    if search_text:
        flv = flv.loc[
            flv["Category"].astype(str).str.lower().str.contains(search_text, na=False)
            | flv["Line Item"].astype(str).str.lower().str.contains(search_text, na=False)
        ]

    sort_map = {
        "Largest Variance": lambda df: df.assign(_s=df["Variance"].abs()).sort_values(["_s", "Line Item"], ascending=[False, True]),
        "Largest Variance %": lambda df: df.assign(_s=df["Variance %"].fillna(-1)).sort_values(["_s", "Line Item"], ascending=[False, True]),
        "Largest Planned": lambda df: df.sort_values(["Planned", "Line Item"], ascending=[False, True]),
        "Largest Actual": lambda df: df.sort_values(["Actual", "Line Item"], ascending=[False, True]),
        "A-Z": lambda df: df.sort_values(["Section", "Category", "Line Item"]),
    }
    flv = sort_map.get(sort_order, lambda df: df)(flv)
    if "_s" in flv.columns:
        flv = flv.drop(columns="_s")

    if flv.empty:
        st.info("No line items match the selected filters.")
        return

    status_summary = flv["Status"].value_counts()
    m1, m2, m3 = st.columns(3)
    m1.metric("Matched Summary Rows", int(status_summary.get("Matched", 0)))
    m2.metric("Planned Only", int(status_summary.get("Planned only", 0)))
    m3.metric("Actual Only", int(status_summary.get("Actual only", 0)))
    st.caption(f"Selected-month variance · Rows shown: {len(flv):,}")
    render_simple_table(flv, currency_columns=["Planned", "Actual", "Variance"], percent_columns=["Variance %"])

    line_exceptions = flv.loc[
        (flv["Status"] != "Matched") | (flv["Variance %"].fillna(0) > 0.10),
        ["Section", "Category", "Line Item", "Status", "Variance", "Variance %"],
    ]
    if not line_exceptions.empty:
        st.caption("Summary-level exceptions: unmatched or >10% variance")
        render_simple_table(
            line_exceptions,
            currency_columns=["Variance"],
            percent_columns=["Variance %"],
            enable_sort=True,
            sort_key_prefix="avb_line_exceptions",
            sortable_columns=["Section", "Category", "Line Item", "Status", "Variance", "Variance %"],
        )


def render_weekly_drilldown(week_meta: list[dict[str, Any]]) -> None:
    st.markdown("**Weekly Drill-Down**")
    for week in week_meta:
        with st.expander(week["label"], expanded=False):
            dl_col, _ = st.columns([1, 3])
            with dl_col:
                week_export = build_week_raw_export(week)
                slug = week["key"].lower().replace(" ", "_")
                st.download_button(
                    "Download Raw Week Data (Excel)",
                    data=week_export,
                    file_name=f"{slug}_raw_data.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key=f"download_{slug}",
                    use_container_width=True,
                )
            if week.get("holidays"):
                holiday_text = "; ".join(
                    f"{full_date_with_day(e['date'])}: {e['name']}" for e in week["holidays"]
                )
                st.markdown(
                    f"""
                    <div style="border:1px solid #fecaca; background:#fef2f2; color:#991b1b;
                                padding:10px 12px; margin-bottom:12px; border-radius:6px;">
                        <strong>Anticipated bank holidays:</strong> {holiday_text}
                    </div>
                    """,
                    unsafe_allow_html=True,
                )
            inflows = summarize_inflow_detail(week["plan_receipts_detail"], value_column="expected_base", value_label="Expected Collection")
            outflows = summarize_outflow_detail(week["plan_outflows_detail"])
            detail_filter = st.radio(
                "Weekly detail view", options=["Inflows", "Outflows"], index=0,
                horizontal=True,
                key=f"weekly_detail_{week['key'].lower().replace(' ', '_')}",
                label_visibility="collapsed",
            )
            slug = week["key"].lower().replace(" ", "_")
            if detail_filter == "Inflows":
                st.caption("Summary inflows grouped by billing type and client.")
                render_simple_table(inflows, currency_columns=["Expected Collection"], enable_sort=True,
                                    sort_key_prefix=f"{slug}_inflows",
                                    sortable_columns=["Billing Type", "Client", "Due Window", "Forecast Window", "Expected Collection", "Source Rows"])
            else:
                st.caption("Summary outflows grouped by category and vendor.")
                render_simple_table(outflows, currency_columns=["Amount"], enable_sort=True,
                                    sort_key_prefix=f"{slug}_outflows",
                                    sortable_columns=["Category", "Vendor", "Payment Window", "Amount", "Source Rows"])


def render_week_holiday_summary(week_meta: list[dict[str, Any]]) -> None:
    st.caption(
        "Bank-holiday overlay covers the selected region. Second and fourth Saturdays are included. Sundays excluded."
    )
    columns = st.columns(max(len(week_meta), 1))
    for col, week in zip(columns, week_meta):
        with col:
            with st.container(border=True):
                st.markdown(f"**{week['key']}**")
                if week.get("holidays"):
                    holiday_lines = "\n".join(
                        f"- {short_date_label(e['date'])}: {e['name']}" for e in week["holidays"]
                    )
                    st.markdown(holiday_lines)
                else:
                    st.caption("No anticipated bank holidays")


# ══════════════════════════════════════════════════════════════════════════════
#  EXCEL EXPORT BUILDERS
# ══════════════════════════════════════════════════════════════════════════════

def _auto_column_widths(sheet) -> None:
    for col_cells in sheet.columns:
        max_length = max((len(str(cell.value or "")) for cell in col_cells), default=0)
        sheet.column_dimensions[col_cells[0].column_letter].width = min(max_length + 2, 42)


def _get_commentary(key: str) -> str:
    return st.session_state.get(f"commentary_{key}", "")


def _excel_compatible_value(value: Any) -> Any:
    if value is None or pd.isna(value):
        return None
    if isinstance(value, pd.Timestamp):
        return value.to_pydatetime()
    return value


def _build_unique_sheet_title(desired_title: str, existing_titles: list[str]) -> str:
    cleaned = re.sub(r"[\[\]\*\?/\\:]", " ", str(desired_title or "Sheet")).strip() or "Sheet"
    cleaned = cleaned[:31]
    existing_lookup = {title.lower() for title in existing_titles}
    candidate = cleaned
    suffix = 2
    while candidate.lower() in existing_lookup:
        suffix_text = f" ({suffix})"
        candidate = f"{cleaned[: max(0, 31 - len(suffix_text))]}{suffix_text}"
        suffix += 1
    return candidate


def _get_header_row_index(worksheet) -> int:
    for row_idx in range(1, worksheet.max_row + 1):
        if any(worksheet.cell(row=row_idx, column=col_idx).value not in (None, "") for col_idx in range(1, worksheet.max_column + 1)):
            return row_idx
    return 1


def _write_dataframe_to_sheet(worksheet, df: pd.DataFrame, header_row: int | None = None) -> None:
    header_row = header_row or _get_header_row_index(worksheet)
    max_column = max(worksheet.max_column, len(df.columns), 1)
    for row_idx in range(header_row, worksheet.max_row + 1):
        for col_idx in range(1, max_column + 1):
            worksheet.cell(row=row_idx, column=col_idx).value = None

    for col_idx, column_name in enumerate(df.columns, start=1):
        worksheet.cell(row=header_row, column=col_idx).value = column_name

    for row_offset, row_values in enumerate(df.itertuples(index=False, name=None), start=1):
        for col_idx, value in enumerate(row_values, start=1):
            worksheet.cell(row=header_row + row_offset, column=col_idx).value = _excel_compatible_value(value)

    worksheet.freeze_panes = f"A{header_row + 1}"
    _auto_column_widths(worksheet)


def _extract_receivable_reference(description: Any, counterparty: Any) -> str:
    description_text = str(description or "").strip()
    counterparty_text = str(counterparty or "").strip()
    if " | " not in description_text:
        return ""
    left_part, right_part = description_text.split(" | ", 1)
    if counterparty_text and standardize_label(left_part) != standardize_label(counterparty_text):
        return ""
    return right_part.strip()


def build_receivables_source_sheet(receivables_df: pd.DataFrame, sheet_meta: dict[str, Any] | None = None) -> pd.DataFrame:
    sheet_meta = sheet_meta or {}
    columns = list(sheet_meta.get("columns") or DEFAULT_RECEIVABLE_EXPORT_COLUMNS)
    column_map = dict(sheet_meta.get("column_map") or detect_receivables_columns(columns))
    if not columns or not column_map.get("amount"):
        columns = DEFAULT_RECEIVABLE_EXPORT_COLUMNS.copy()
        column_map = detect_receivables_columns(columns)

    export_df = pd.DataFrame({column: [None] * len(receivables_df) for column in columns})
    if receivables_df.empty:
        return export_df

    invoice_date_series = receivables_df.get("invoice_date", pd.Series(pd.NaT, index=receivables_df.index))
    amount_series = receivables_df.get("amount", pd.Series(0.0, index=receivables_df.index))
    counterparty_series = receivables_df.get("counterparty", pd.Series("", index=receivables_df.index))
    description_series = receivables_df.get("description", pd.Series("", index=receivables_df.index))
    ageing_series = receivables_df.get("ageing", pd.Series(0.0, index=receivables_df.index))
    tentative_series = receivables_df.get("tentative_collection_date", pd.Series(pd.NaT, index=receivables_df.index))
    due_date_series = receivables_df.get("due_date", pd.Series(pd.NaT, index=receivables_df.index))
    billing_series = receivables_df.get("billing_type", pd.Series("", index=receivables_df.index))
    if billing_series.fillna("").astype(str).str.strip().eq("").all():
        billing_series = receivables_df.get("billing_bucket", pd.Series("", index=receivables_df.index))

    if column_map.get("date"):
        export_df[column_map["date"]] = pd.to_datetime(invoice_date_series, errors="coerce")
    if column_map.get("amount"):
        export_df[column_map["amount"]] = pd.to_numeric(amount_series, errors="coerce")
    if column_map.get("client"):
        export_df[column_map["client"]] = counterparty_series
    if column_map.get("reference"):
        export_df[column_map["reference"]] = [
            _extract_receivable_reference(desc, party)
            for desc, party in zip(description_series, counterparty_series)
        ]
    if column_map.get("ageing"):
        export_df[column_map["ageing"]] = pd.to_numeric(ageing_series, errors="coerce")
    if column_map.get("tentative_date"):
        export_df[column_map["tentative_date"]] = pd.to_datetime(tentative_series, errors="coerce")
    if column_map.get("due_date"):
        export_df[column_map["due_date"]] = pd.to_datetime(due_date_series, errors="coerce")
    if column_map.get("billing_type"):
        export_df[column_map["billing_type"]] = billing_series.fillna("").astype(str)
    return export_df


def build_outflow_source_sheet(
    outflow_df: pd.DataFrame,
    sheet_meta: dict[str, Any] | None = None,
) -> pd.DataFrame:
    sheet_meta = sheet_meta or {}
    columns = list(sheet_meta.get("columns") or DEFAULT_OUTFLOW_EXPORT_COLUMNS)
    column_map = dict(sheet_meta.get("column_map") or detect_outflow_columns(columns))
    if not columns or not column_map.get("amount") or not (column_map.get("planned_date") or column_map.get("actual_date")):
        columns = DEFAULT_OUTFLOW_EXPORT_COLUMNS.copy()
        column_map = detect_outflow_columns(columns)

    export_df = pd.DataFrame({column: [None] * len(outflow_df) for column in columns})
    if outflow_df.empty:
        return export_df

    date_series = outflow_df.get("date", pd.Series(pd.NaT, index=outflow_df.index))
    amount_series = outflow_df.get("amount", pd.Series(0.0, index=outflow_df.index))
    description_series = outflow_df.get("description", pd.Series("", index=outflow_df.index))
    vendor_series = outflow_df.get("vendor_name", pd.Series("", index=outflow_df.index))
    date_target = column_map.get("planned_date") or column_map.get("actual_date")
    if date_target:
        export_df[date_target] = pd.to_datetime(date_series, errors="coerce")
    if column_map.get("amount"):
        export_df[column_map["amount"]] = pd.to_numeric(amount_series, errors="coerce")
    if column_map.get("description"):
        export_df[column_map["description"]] = description_series.fillna("").astype(str)
    if column_map.get("vendor"):
        export_df[column_map["vendor"]] = vendor_series.fillna("").astype(str)
    return export_df


def build_reuploadable_source_export(
    source: Any,
    receivables_df: pd.DataFrame,
    outflows_df: pd.DataFrame,
    workbook_layout: dict[str, Any] | None = None,
) -> bytes:
    workbook_layout = workbook_layout or {}
    sheet_meta_map = workbook_layout.get("sheet_meta", {})

    reset_source_pointer(source)
    workbook = load_workbook(source)
    reset_source_pointer(source)
    receivables_export_df = receivables_df.copy()

    existing_inflow_sheets = workbook_layout.get("inflow_sheet_names", [])
    inflow_sheet_lookup = {
        standardize_label(sheet_name): sheet_name
        for sheet_name in existing_inflow_sheets
        if str(sheet_name).strip()
    }
    fallback_inflow_sheet_name = existing_inflow_sheets[0] if existing_inflow_sheets else "Receivables"
    if not receivables_export_df.empty:
        if "source_sheet" not in receivables_export_df.columns:
            receivables_export_df["source_sheet"] = fallback_inflow_sheet_name
        else:
            receivables_export_df["source_sheet"] = (
                receivables_export_df["source_sheet"]
                .fillna("")
                .astype(str)
                .str.strip()
                .replace("", fallback_inflow_sheet_name)
            )

    def inflow_rows_for(sheet_name: str) -> pd.DataFrame:
        if receivables_export_df.empty:
            return receivables_export_df.iloc[0:0].copy()
        if "source_sheet" not in receivables_export_df.columns:
            return receivables_export_df.reset_index(drop=True)
        return receivables_export_df.loc[
            receivables_export_df["source_sheet"].fillna("").astype(str).map(standardize_label) == standardize_label(sheet_name)
        ].reset_index(drop=True)

    for sheet_name in existing_inflow_sheets:
        if sheet_name not in workbook.sheetnames:
            workbook.create_sheet(sheet_name)
        worksheet = workbook[sheet_name]
        receivables_export = build_receivables_source_sheet(
            inflow_rows_for(sheet_name),
            sheet_meta=sheet_meta_map.get(sheet_name),
        )
        _write_dataframe_to_sheet(worksheet, receivables_export)

    if receivables_export_df.empty:
        additional_inflow_sheets: list[str] = []
    else:
        additional_inflow_sheets = [
            sheet_name
            for sheet_name in order_sheet_names(receivables_export_df.get("source_sheet", pd.Series(dtype="object")).dropna().astype(str).tolist())
            if standardize_label(sheet_name) not in inflow_sheet_lookup
        ]

    seen_new_inflow_sheets: set[str] = set()
    for sheet_name in additional_inflow_sheets:
        sheet_key = standardize_label(sheet_name)
        if not sheet_key or sheet_key in seen_new_inflow_sheets:
            continue
        seen_new_inflow_sheets.add(sheet_key)
        worksheet = workbook.create_sheet(_build_unique_sheet_title(sheet_name, workbook.sheetnames))
        receivables_export = build_receivables_source_sheet(inflow_rows_for(sheet_name))
        _write_dataframe_to_sheet(worksheet, receivables_export)

    existing_outflow_sheets = workbook_layout.get("outflow_sheet_names", [])
    existing_outflow_lookup = {
        standardize_label(sheet_name): sheet_name
        for sheet_name in existing_outflow_sheets
        if str(sheet_name).strip()
    }

    def outflow_rows_for(sheet_name: str) -> pd.DataFrame:
        if outflows_df.empty:
            return outflows_df.iloc[0:0].copy()
        return outflows_df.loc[
            outflows_df["sheet_name"].fillna("").astype(str).map(standardize_label) == standardize_label(sheet_name)
        ].reset_index(drop=True)

    for sheet_name in existing_outflow_sheets:
        if sheet_name not in workbook.sheetnames:
            workbook.create_sheet(sheet_name)
        worksheet = workbook[sheet_name]
        export_df = build_outflow_source_sheet(
            outflow_rows_for(sheet_name),
            sheet_meta=sheet_meta_map.get(sheet_name),
        )
        _write_dataframe_to_sheet(worksheet, export_df)

    if outflows_df.empty:
        additional_categories: list[str] = []
    else:
        additional_categories = [
            category
            for category in order_outflow_sheets(outflows_df["sheet_name"].dropna().astype(str).tolist())
            if standardize_label(category) not in existing_outflow_lookup
        ]

    seen_new_categories: set[str] = set()
    for category in additional_categories:
        category_key = standardize_label(category)
        if not category_key or category_key in seen_new_categories:
            continue
        seen_new_categories.add(category_key)
        sheet_title = _build_unique_sheet_title(category, workbook.sheetnames)
        worksheet = workbook.create_sheet(sheet_title)
        export_df = build_outflow_source_sheet(outflow_rows_for(category))
        _write_dataframe_to_sheet(worksheet, export_df)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output.getvalue()


def build_weekly_breakdown_export(
    weekly_cashflow: dict[str, Any],
    selected_month: pd.Timestamp,
) -> bytes:
    output = BytesIO()
    weekly_table = round_numeric_columns(weekly_cashflow["weekly_table"].copy())

    holiday_rows = []
    for week in weekly_cashflow["week_meta"]:
        if week.get("holidays"):
            for e in week["holidays"]:
                holiday_rows.append({"Week": week["key"], "Week Range": week["label"],
                                     "Holiday Date": format_date(e["date"]), "Holiday Name": e["name"]})
        else:
            holiday_rows.append({"Week": week["key"], "Week Range": week["label"],
                                 "Holiday Date": "", "Holiday Name": "No anticipated bank holidays"})
    holiday_df = pd.DataFrame(holiday_rows)

    # Commentary
    exec_commentary = _get_commentary("executive_summary")
    avb_commentary = _get_commentary("avb_variance")
    commentary_df = pd.DataFrame([
        {"Section": "Executive Summary", "Commentary": exec_commentary},
        {"Section": "Actual vs Budget", "Commentary": avb_commentary},
    ])

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        weekly_table.to_excel(writer, index=False, sheet_name="Weekly Breakdown")
        format_date_columns(holiday_df).to_excel(writer, index=False, sheet_name="Week Holidays")
        commentary_df.to_excel(writer, index=False, sheet_name="CFO Commentary")

        ws = writer.sheets["Weekly Breakdown"]
        ws.freeze_panes = "B2"
        apply_excel_number_formats(ws)
        _auto_column_widths(ws)
        _auto_column_widths(writer.sheets["Week Holidays"])
        _auto_column_widths(writer.sheets["CFO Commentary"])

    output.seek(0)
    return output.getvalue()


def build_week_raw_export(week: dict[str, Any]) -> bytes:
    output = BytesIO()
    inflows = prepare_inflow_line_item_table(week["plan_receipts_detail"])
    outflows = prepare_outflow_line_item_table(week["plan_outflows_detail"])
    inflow_summary = round_numeric_columns(
        summarize_inflow_detail(week["plan_receipts_detail"], value_column="expected_base", value_label="Expected Collection")
    )
    outflow_summary = round_numeric_columns(summarize_outflow_detail(week["plan_outflows_detail"]))

    holiday_rows = [
        {"Week": week["key"], "Week Range": week["label"],
         "Holiday Date": format_date(e["date"]), "Holiday Name": e["name"]}
        for e in week.get("holidays", [])
    ]
    holiday_df = (
        pd.DataFrame(holiday_rows)
        if holiday_rows
        else pd.DataFrame([{"Week": week["key"], "Week Range": week["label"],
                            "Holiday Date": "", "Holiday Name": "No anticipated bank holidays"}])
    )

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        inflow_summary.to_excel(writer, index=False, sheet_name="Inflows Summary")
        outflow_summary.to_excel(writer, index=False, sheet_name="Outflows Summary")
        inflows.to_excel(writer, index=False, sheet_name="Inflows Raw")
        outflows.to_excel(writer, index=False, sheet_name="Outflows Raw")
        format_date_columns(holiday_df).to_excel(writer, index=False, sheet_name="Holiday Overlay")
        for sheet in writer.sheets.values():
            sheet.freeze_panes = "A2"
            apply_excel_number_formats(sheet)
            _auto_column_widths(sheet)

    output.seek(0)
    return output.getvalue()


def build_month_raw_export(weekly_cashflow: dict[str, Any], selected_month: pd.Timestamp) -> bytes:
    output = BytesIO()
    month_start = weekly_cashflow["selected_month_start"]
    month_end = weekly_cashflow["selected_month_end"]
    period_start = weekly_cashflow.get("selected_period_start", month_start)

    inflow_frames, outflow_frames, holiday_rows = [], [], []
    for week in weekly_cashflow["week_meta"]:
        if not week["plan_receipts_detail"].empty:
            iw = week["plan_receipts_detail"].copy()
            iw["Week"] = week["key"]
            iw["Week Range"] = week["label"]
            inflow_frames.append(iw)
        if not week["plan_outflows_detail"].empty:
            ow = week["plan_outflows_detail"].copy()
            ow["Week"] = week["key"]
            ow["Week Range"] = week["label"]
            outflow_frames.append(ow)
        if week.get("holidays"):
            for e in week["holidays"]:
                holiday_rows.append({"Week": week["key"], "Week Range": week["label"],
                                     "Holiday Date": format_date(e["date"]), "Holiday Name": e["name"]})

    inflows = pd.concat(inflow_frames, ignore_index=True) if inflow_frames else pd.DataFrame()
    outflows = pd.concat(outflow_frames, ignore_index=True) if outflow_frames else pd.DataFrame()
    inflows = round_numeric_columns(inflows, percent_columns=["base_probability", "best_probability", "worst_probability"]) if not inflows.empty else inflows
    outflows = round_numeric_columns(outflows) if not outflows.empty else outflows

    all_inflows = pd.concat(inflow_frames, ignore_index=True) if inflow_frames else pd.DataFrame()
    all_outflows = pd.concat(outflow_frames, ignore_index=True) if outflow_frames else pd.DataFrame()
    inflow_summary = round_numeric_columns(
        summarize_inflow_detail(all_inflows, value_column="expected_base", value_label="Expected Collection")
    )
    outflow_summary = round_numeric_columns(summarize_outflow_detail(all_outflows))

    keep_inflow_cols = ["Week", "Week Range", "invoice_date", "tentative_collection_date",
                        "forecast_collection_date", "due_date", "billing_type", "billing_bucket",
                        "counterparty", "description", "amount", "aging_bucket",
                        "base_probability", "expected_base"]
    if not inflows.empty:
        inflows = inflows[[c for c in keep_inflow_cols if c in inflows.columns]]
    else:
        inflows = pd.DataFrame(columns=keep_inflow_cols)

    if not outflows.empty:
        outflows = outflows[[c for c in ["Week", "Week Range", "sheet_name", "vendor_name", "description", "date", "amount"] if c in outflows.columns]].rename(columns={
            "sheet_name": "Category", "vendor_name": "Vendor", "description": "Particular", "date": "Payment Date", "amount": "Amount"
        })
    else:
        outflows = pd.DataFrame(columns=["Week", "Week Range", "Category", "Vendor", "Particular", "Payment Date", "Amount"])

    holiday_df = (
        pd.DataFrame(holiday_rows)
        if holiday_rows
        else pd.DataFrame([{"Week": "", "Week Range": "", "Holiday Date": "", "Holiday Name": "No anticipated bank holidays"}])
    )

    cover_df = pd.DataFrame([
        {"Field": "Selected Month", "Value": month_label(selected_month)},
        {"Field": "Balance As Of Date", "Value": format_date(period_start)},
        {"Field": "Month Start", "Value": format_date(month_start)},
        {"Field": "Month End", "Value": format_date(month_end)},
        {"Field": "Weeks in Export", "Value": len(weekly_cashflow["week_meta"])},
        {"Field": "CFO Commentary", "Value": _get_commentary("executive_summary")},
    ])

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        format_date_columns(cover_df).to_excel(writer, index=False, sheet_name="Summary")
        inflow_summary.to_excel(writer, index=False, sheet_name="Inflows Summary")
        outflow_summary.to_excel(writer, index=False, sheet_name="Outflows Summary")
        inflows.to_excel(writer, index=False, sheet_name="Inflows Raw")
        outflows.to_excel(writer, index=False, sheet_name="Outflows Raw")
        format_date_columns(holiday_df).to_excel(writer, index=False, sheet_name="Holiday Overlay")
        for sheet in writer.sheets.values():
            sheet.freeze_panes = "A2"
            apply_excel_number_formats(sheet)
            _auto_column_widths(sheet)

    output.seek(0)
    return output.getvalue()


# ══════════════════════════════════════════════════════════════════════════════
#  STYLES & LAYOUT
# ══════════════════════════════════════════════════════════════════════════════

def inject_styles() -> None:
    st.markdown(
        """
        <style>
        .stApp { background: #f7f7f5; color: #111827; }
        .block-container { padding-top: 1.5rem; padding-bottom: 2.0rem; max-width: 1460px; }
        div[data-testid="metric-container"] {
            border: 1px solid #e5e7eb; padding: 12px 14px; border-radius: 6px;
            background: #ffffff; box-shadow: 0 1px 2px rgba(17,24,39,0.04);
        }
        div[data-testid="metric-container"] label { font-size: 0.82rem; }
        [data-testid="stSidebar"] { background: #ffffff; border-right: 1px solid #e5e7eb; }
        .app-banner {
            background: #ffffff; border: 1px solid #e5e7eb;
            border-left: 4px solid #111827; padding: 18px 20px; margin-bottom: 18px;
        }
        .app-banner-title { font-size: 1.05rem; font-weight: 700; margin-bottom: 6px; }
        .app-banner-subtitle { font-size: 0.93rem; color: #4b5563; }
        .context-grid {
            display: grid; grid-template-columns: repeat(4, minmax(0, 1fr));
            gap: 10px; margin-top: 14px;
        }
        .context-card {
            background: #fafaf9; border: 1px solid #e7e5e4;
            padding: 12px 14px; border-radius: 6px;
        }
        .context-label {
            font-size: 0.75rem; color: #6b7280;
            text-transform: uppercase; letter-spacing: 0.03em; margin-bottom: 3px;
        }
        .context-value { font-size: 0.95rem; font-weight: 600; color: #111827; }
        .section-header { margin-top: 18px; margin-bottom: 12px; }
        .section-title { font-size: 1.05rem; font-weight: 700; color: #111827; margin-bottom: 2px; }
        .section-subtitle { font-size: 0.90rem; color: #6b7280; }
        </style>
        """,
        unsafe_allow_html=True,
    )


def default_month_index(available_months: list[pd.Timestamp]) -> int:
    current_month = build_month_start(TODAY)
    for idx, month in enumerate(available_months):
        if build_month_start(month) == current_month:
            return idx
    return 0


def render_section_header(title: str, subtitle: str) -> None:
    st.markdown(
        f"""
        <div class="section-header">
            <div class="section-title">{title}</div>
            <div class="section-subtitle">{subtitle}</div>
        </div>
        """,
        unsafe_allow_html=True,
    )


def render_app_banner(
    selected_month: pd.Timestamp,
    balance_as_of_date: pd.Timestamp,
    planned_source_name: str,
    outflow_categories: int,
    holiday_region: str,
    actual_file_name: str | None,
    load_timestamp: str | None = None,
    tds_rate: float = 0.0,
    min_cash_threshold: float = 0.0,
) -> None:
    actual_status = actual_file_name if actual_file_name else "Not loaded"
    load_ts_display = load_timestamp if load_timestamp else "—"
    tds_display = f"{tds_rate*100:.1f}%" if tds_rate > 0 else "None"
    floor_display = format_currency(min_cash_threshold) if min_cash_threshold > 0 else "—"
    st.markdown(
        f"""
        <div class="app-banner">
            <div class="app-banner-title">Liquidity review workspace for finance leadership</div>
            <div class="app-banner-subtitle">
                Overview first, then weekly cashflow detail, collection scenarios, roll-forward, and actual-vs-budget validation.
            </div>
            <div class="context-grid">
                <div class="context-card">
                    <div class="context-label">Selected Month</div>
                    <div class="context-value">{month_label(selected_month)}</div>
                </div>
                <div class="context-card">
                    <div class="context-label">Balance As Of</div>
                    <div class="context-value">{format_date(balance_as_of_date)}</div>
                </div>
                <div class="context-card">
                    <div class="context-label">Planned Workbook</div>
                    <div class="context-value">{planned_source_name}</div>
                </div>
                <div class="context-card">
                    <div class="context-label">Outflow Categories</div>
                    <div class="context-value">{outflow_categories}</div>
                </div>
                <div class="context-card">
                    <div class="context-label">Holiday Region</div>
                    <div class="context-value">{holiday_region}</div>
                </div>
                <div class="context-card">
                    <div class="context-label">Actuals Status</div>
                    <div class="context-value">{actual_status}</div>
                </div>
                <div class="context-card">
                    <div class="context-label">TDS Rate Applied</div>
                    <div class="context-value">{tds_display}</div>
                </div>
                <div class="context-card">
                    <div class="context-label">Cash Floor</div>
                    <div class="context-value">{floor_display}</div>
                </div>
            </div>
            <div style="margin-top:10px; font-size:0.78rem; color:#9ca3af;">
                Workbook loaded: {load_ts_display}
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )


def get_sample_sheet_names() -> list[str]:
    if not DEFAULT_SOURCE.exists():
        return []
    workbook = load_workbook(DEFAULT_SOURCE, read_only=True, data_only=False)
    return workbook.sheetnames


def render_upload_prompt() -> None:
    sample_sheets = get_sample_sheet_names()
    sheet_list = "".join(f"<li>{s}</li>" for s in sample_sheets) if sample_sheets else "<li>Receivables sheet</li><li>Separate outflow sheets by category</li>"
    st.title(APP_TITLE)
    st.caption("Upload today's workbook in the approved format to start the liquidity model.")
    st.markdown(
        f"""
        <div class="app-banner">
            <div class="app-banner-title">Planned workbook required</div>
            <div class="app-banner-subtitle">The app runs only on the workbook uploaded for the day. You can add or remove outflow sheets as needed, as long as the key headers stay recognizable.</div>
            <div style="margin-top:14px; font-size:0.94rem;">
                <strong>Expected format reference</strong>
                <ul style="margin-top:8px;">{sheet_list}</ul>
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )


# ══════════════════════════════════════════════════════════════════════════════
#  UPLOAD HANDLERS
# ══════════════════════════════════════════════════════════════════════════════

def handle_planned_upload_change() -> None:
    uploaded = st.session_state.get("planned_uploader_widget")
    st.session_state["planned_workbook_file"] = uploaded
    load_time = datetime.now().strftime("%d-%m-%Y %H:%M:%S") if uploaded else None
    st.session_state["planned_load_time"] = load_time
    persist_uploaded_file("planned", uploaded, load_time=load_time)
    st.session_state.pop("selected_forecast_month", None)
    # Clear any in-app edits so fresh data seeds from the new file
    st.session_state.pop("_edit_source_hash", None)
    st.session_state.pop("edited_receivables", None)
    st.session_state.pop("edited_outflows", None)


def handle_actual_upload_change() -> None:
    uploaded = st.session_state.get("actual_uploader_widget")
    st.session_state["actual_workbook_file"] = uploaded
    persist_uploaded_file("actual", uploaded)


# ══════════════════════════════════════════════════════════════════════════════
#  MAIN APPLICATION
# ══════════════════════════════════════════════════════════════════════════════

def main() -> None:
    st.set_page_config(page_title=APP_TITLE, layout="wide", initial_sidebar_state="expanded")
    inject_styles()
    navigation_options = [
        "Executive Summary",
        "Group Cashflow",
        "Collection Engine",
        "3-Month Cash Flow View",
        "Actual vs Budget",
        "✏️ Edit Data",
    ]
    requested_view = get_requested_view(
        GROUP_CASHFLOW_VIEW if has_weekly_drilldown_query() else DEFAULT_NAV_VIEW
    )
    if requested_view in navigation_options and "selected_view" not in st.session_state:
        st.session_state["selected_view"] = requested_view

    # ── Sidebar ───────────────────────────────────────────────────────────────
    with st.sidebar:
        st.title(APP_TITLE)
        st.caption("CFO-level liquidity control room")
        st.markdown("---")
        st.markdown("**Navigation**")
        selected_view = st.radio(
            "Select Section",
            navigation_options,
            label_visibility="collapsed",
            key="selected_view",
        )
        st.markdown("---")

        # ── Data Sources ──────────────────────────────────────────────────────
        st.markdown("**Data Sources**")
        st.file_uploader(
            "Planned workbook (Excel)", type=["xlsx", "xls"],
            key="planned_uploader_widget", on_change=handle_planned_upload_change,
        )
        st.file_uploader(
            "Actual workbook (Excel)", type=["xlsx", "xls"],
            key="actual_uploader_widget", on_change=handle_actual_upload_change,
        )
        st.markdown("---")

    # ── Load planned data ────────────────────────────────────────────────────
    planned_upload = st.session_state.get("planned_workbook_file", st.session_state.get("planned_uploader_widget"))
    if planned_upload is None and has_weekly_drilldown_query():
        cached_planned, planned_meta = load_cached_upload("planned")
        if cached_planned is not None:
            planned_upload = cached_planned
            st.session_state["planned_workbook_file"] = cached_planned
            if planned_meta.get("load_time"):
                st.session_state["planned_load_time"] = planned_meta["load_time"]
    if planned_upload is None:
        render_upload_prompt()
        return
    planned_source = planned_upload

    try:
        plan_data = load_data(planned_source, actual_mode=False)
    except Exception as exc:
        st.error(f"Unable to load planned workbook: {exc}")
        return

    # ── Seed / reconcile in-app edit state ──────────────────────────────────
    # On first load (or when file changes) copy raw Excel data into session
    # state.  Subsequent reruns use the session-state copy so in-app edits
    # survive without re-uploading.
    _current_hash = file_md5(planned_upload)
    if st.session_state.get("_edit_source_hash") != _current_hash:
        st.session_state["_edit_source_hash"] = _current_hash
        initialize_edit_history("receivables", plan_data["receivables"])
        initialize_edit_history("outflows", plan_data["outflows"])
    else:
        for dataset_key, fallback_df in (
            ("receivables", plan_data["receivables"]),
            ("outflows", plan_data["outflows"]),
        ):
            if (
                _edit_history_key(dataset_key) not in st.session_state
                or _edit_history_index_key(dataset_key) not in st.session_state
                or _edit_history_log_key(dataset_key) not in st.session_state
            ):
                initialize_edit_history(
                    dataset_key,
                    st.session_state.get(_edit_state_key(dataset_key), fallback_df).copy(),
                )

    # Keep original Excel data available (used for reset in Edit Data section)
    _excel_receivables = plan_data["receivables"].copy()
    _excel_outflows = plan_data["outflows"].copy()

    # Override plan_data with session-state (possibly user-edited) versions
    plan_data["receivables"] = st.session_state["edited_receivables"].copy()
    plan_data["outflows"] = st.session_state["edited_outflows"].copy()
    # Keep outflow_sheet_order in sync with edited outflows
    if not plan_data["outflows"].empty:
        _edited_sheets = plan_data["outflows"]["sheet_name"].dropna().unique().tolist()
        plan_data["outflow_sheet_order"] = order_outflow_sheets(_edited_sheets)

    available_months = plan_data["available_months"]
    if not available_months:
        st.error("No valid forecast dates found in the planned workbook.")
        return

    # ── Sidebar: Forecast Settings ─────────────────────────────────────────
    with st.sidebar:
        if planned_upload:
            load_time = st.session_state.get("planned_load_time", "—")
            st.success(f"Planned: {planned_upload.name}")
            st.caption(f"Loaded: {load_time}")

        actual_upload = st.session_state.get("actual_workbook_file", st.session_state.get("actual_uploader_widget"))
        if actual_upload:
            st.success(f"Actual: {actual_upload.name}")
        else:
            st.caption("Upload an actual workbook to enable variance tracking.")

        st.markdown("**Forecast Settings**")
        requested_forecast_month = get_query_params().get(FORECAST_MONTH_QUERY_KEY, "").strip()
        if requested_forecast_month and st.session_state.get("selected_forecast_month") is None:
            requested_month_start = build_month_start(requested_forecast_month)
            matched_month = next(
                (month for month in available_months if build_month_start(month) == requested_month_start),
                None,
            )
            if matched_month is not None:
                st.session_state["selected_forecast_month"] = matched_month
        selected_month = st.selectbox(
            "Forecast Month",
            options=available_months,
            format_func=month_label,
            index=default_month_index(available_months),
            key="selected_forecast_month",
        )
        sync_forecast_month_query_param(selected_month)
        current_bank_balance = st.number_input(
            "Current Bank Balance (₹)", min_value=0.0, value=5_000_000.0,
            step=100_000.0, format="%.0f",
        )
        holiday_region = st.selectbox(
            "Holiday Region", options=list(STATE_BANK_HOLIDAYS.keys()), index=0,
            help="Used for week-level bank-holiday highlighting.",
        )

        month_start = build_month_start(selected_month)
        month_end = build_month_end(selected_month)
        balance_date_key = "bank_balance_as_of_date"
        requested_balance_date = get_query_params().get(BALANCE_DATE_QUERY_KEY, "").strip()
        if requested_balance_date:
            requested_balance_ts = pd.Timestamp(requested_balance_date).normalize()
            if month_start <= requested_balance_ts <= month_end:
                requested_balance_value = requested_balance_ts.date()
                if st.session_state.get(balance_date_key) != requested_balance_value:
                    st.session_state[balance_date_key] = requested_balance_value
        stored = st.session_state.get(balance_date_key)
        if stored is None:
            st.session_state[balance_date_key] = default_balance_as_of_date(selected_month).date()
        else:
            ts = pd.Timestamp(stored).normalize()
            if ts < month_start or ts > month_end:
                st.session_state[balance_date_key] = default_balance_as_of_date(selected_month).date()

        balance_as_of_input = st.date_input(
            "Bank Balance As Of Date",
            min_value=month_start.date(), max_value=month_end.date(),
            key=balance_date_key, format="DD-MM-YYYY",
            help="Forecast starts from this date.",
        )
        balance_as_of_date = pd.Timestamp(balance_as_of_input).normalize()
        st.markdown("---")

        # ── Model Assumptions ─────────────────────────────────────────────────
        st.markdown("**Model Assumptions**")

        # Minimum cash floor
        min_cash_threshold = st.number_input(
            "Minimum Cash Floor (₹)",
            min_value=0.0, value=DEFAULT_MIN_CASH_THRESHOLD,
            step=100_000.0, format="%.0f",
            help="Alert triggers if projected closing cash falls below this level.",
        )

        # TDS rate
        tds_pct = st.slider(
            "TDS Rate on Collections (%)", min_value=0.0, max_value=20.0,
            value=DEFAULT_TDS_RATE * 100, step=0.5,
            help="Percentage deducted at source on gross collection amounts. Net cash = Gross × (1 − TDS%).",
        )
        tds_rate = tds_pct / 100.0

        # Per-bucket collection probabilities
        st.markdown("**Collection Probabilities**")
        st.caption("Base = expected case. Worst = stress case. Adjust per your counterparty mix.")
        with st.expander("0–30 day bucket", expanded=False):
            b0_base = st.slider("Base %", 0, 100, int(DEFAULT_BASE_COLLECTION_PROBABILITIES["0-30"] * 100), key="prob_base_030") / 100
            b0_worst = st.slider("Worst %", 0, 100, int(DEFAULT_WORST_COLLECTION_PROBABILITIES["0-30"] * 100), key="prob_worst_030") / 100
        with st.expander("30–60 day bucket", expanded=False):
            b1_base = st.slider("Base %", 0, 100, int(DEFAULT_BASE_COLLECTION_PROBABILITIES["30-60"] * 100), key="prob_base_3060") / 100
            b1_worst = st.slider("Worst %", 0, 100, int(DEFAULT_WORST_COLLECTION_PROBABILITIES["30-60"] * 100), key="prob_worst_3060") / 100
        with st.expander("60+ day bucket", expanded=False):
            b2_base = st.slider("Base %", 0, 100, int(DEFAULT_BASE_COLLECTION_PROBABILITIES["60+"] * 100), key="prob_base_60p") / 100
            b2_worst = st.slider("Worst %", 0, 100, int(DEFAULT_WORST_COLLECTION_PROBABILITIES["60+"] * 100), key="prob_worst_60p") / 100

        user_base_probs = {"0-30": b0_base, "30-60": b1_base, "60+": b2_base}
        user_worst_probs = {"0-30": b0_worst, "30-60": b1_worst, "60+": b2_worst}
        st.markdown("---")

        # ── OD / Credit Line ──────────────────────────────────────────────────
        st.markdown("**OD / Credit Line**")
        od_sanctioned_limit = st.number_input(
            "OD Sanctioned Limit (₹)", min_value=0.0,
            value=DEFAULT_OD_LIMIT, step=500_000.0, format="%.0f",
            help="Total sanctioned overdraft / cash credit limit. Set to 0 if no OD facility.",
        )
        opening_od_utilization = st.number_input(
            "Opening OD Utilization (₹)", min_value=0.0,
            value=DEFAULT_OPENING_OD_UTILIZATION, step=100_000.0, format="%.0f",
            help="Outstanding OD already utilised as of the opening balance date.",
        )
        od_rate_pa = st.number_input(
            "OD Interest Rate (% p.a.)", min_value=0.0, max_value=50.0,
            value=DEFAULT_OD_RATE_PA, step=0.25, format="%.2f",
            help="Annual interest rate on OD drawn. Used to estimate monthly interest cost.",
        )
        opening_od_headroom = (
            max(od_sanctioned_limit - opening_od_utilization, 0.0)
            if od_sanctioned_limit > 0
            else 0.0
        )
        if od_sanctioned_limit > 0 and opening_od_utilization > od_sanctioned_limit:
            st.warning("Opening OD utilization exceeds the sanctioned limit. Additional OD drawdown will be blocked until headroom is restored.")
        st.caption(f"Available OD headroom at start: {format_currency(opening_od_headroom)}")
        st.caption(f"Weekly surplus above the minimum cash floor of {format_currency(min_cash_threshold)} is used to repay OD.")
        st.markdown("---")

        # ── Model Context ─────────────────────────────────────────────────────
        st.markdown("**Model Context**")
        st.caption(f"Outflow sheets: {len(plan_data['outflow_sheet_order'])}")
        st.caption(f"Receivable lines: {len(plan_data['receivables'])}")
        if planned_upload:
            st.caption(f"File: {planned_upload.name}")
        load_time = st.session_state.get("planned_load_time", "—")
        st.caption(f"Loaded at: {load_time}")

    # ── Load actual data ────────────────────────────────────────────────────
    actual_upload = st.session_state.get("actual_workbook_file", st.session_state.get("actual_uploader_widget"))
    if actual_upload is None and has_weekly_drilldown_query():
        cached_actual, _actual_meta = load_cached_upload("actual")
        if cached_actual is not None:
            actual_upload = cached_actual
            st.session_state["actual_workbook_file"] = cached_actual
    actual_data = None
    if actual_upload is not None:
        try:
            actual_data = load_data(actual_upload, actual_mode=True)
        except Exception as exc:
            st.error(f"Unable to load actual workbook: {exc}")
            return

    # ── Core computation ────────────────────────────────────────────────────
    weekly_results = compute_weekly_cashflow(
        data=plan_data,
        selected_month=selected_month,
        opening_balance=current_bank_balance,
        holiday_region=holiday_region,
        balance_as_of_date=balance_as_of_date,
        actual_data=actual_data,
        base_probs=user_base_probs,
        worst_probs=user_worst_probs,
        tds_rate=tds_rate,
        od_sanctioned_limit=od_sanctioned_limit,
        od_interest_rate_pa=od_rate_pa,
        opening_od_utilization=opening_od_utilization,
        od_buffer_balance=min_cash_threshold,
    )

    next_30_end = balance_as_of_date + pd.Timedelta(days=29)
    next_30_day_outflows = (
        filter_between(plan_data["outflows"], "date", balance_as_of_date, next_30_end)["amount"].sum()
        if not plan_data["outflows"].empty
        else 0.0
    )
    scenario_table, overdue_dependency, sixty_plus_dependency = calculate_collection_engine(
        plan_data["receivables"],
        balance_as_of_date,
        current_bank_balance,
        next_30_day_outflows,
        base_probs=user_base_probs,
        worst_probs=user_worst_probs,
        tds_rate=tds_rate,
    )

    planned_source_name = planned_upload.name
    actual_file_name = actual_upload.name if actual_upload else None
    load_ts = st.session_state.get("planned_load_time")

    # ── Page header ─────────────────────────────────────────────────────────
    st.title(APP_TITLE)
    st.caption("Structured liquidity management for weekly control, receivables sensitivity, and month-end cash visibility.")
    render_app_banner(
        selected_month=selected_month,
        balance_as_of_date=balance_as_of_date,
        planned_source_name=planned_source_name,
        outflow_categories=len(plan_data["outflow_sheet_order"]),
        holiday_region=holiday_region,
        actual_file_name=actual_file_name,
        load_timestamp=load_ts,
        tds_rate=tds_rate,
        min_cash_threshold=min_cash_threshold,
    )

    # ══ SECTION 1: Executive Summary ════════════════════════════════════════
    if selected_view == "Executive Summary":
        render_section_header(
            "Section 1: Executive Summary",
            "Month view of opening cash, total inflows, outflows, month-end balance, and liquidity runway.",
        )
        render_kpis(
            current_bank_balance,
            weekly_results,
            scenario_table,
            next_30_day_outflows,
            overdue_dependency,
            sixty_plus_dependency,
            min_cash_threshold=min_cash_threshold,
            od_sanctioned_limit=od_sanctioned_limit,
        )
        st.markdown("---")
        render_cfo_commentary("executive_summary")

    # ══ SECTION 2: Group Cashflow ════════════════════════════════════════════
    if selected_view == "Group Cashflow":
        render_section_header(
            "Section 2: Group Cashflow",
            "Weekly breakdown for the selected month with monthly forecast and actual columns.",
        )
        st.markdown('<div id="weekly-drilldown-target"></div>', unsafe_allow_html=True)
        if resolve_weekly_drilldown_selection(weekly_results, selected_month) is not None:
            render_weekly_drilldown_selection(weekly_results, selected_month)
            st.markdown("---")
        notes_col, dl_col, raw_col = st.columns([3, 1, 1])
        with notes_col:
            week_notes = " | ".join(w["label"] for w in weekly_results["week_meta"])
            st.caption(week_notes)
            st.caption(f"Forecast starts from {format_date(balance_as_of_date)}. TDS {tds_pct:.1f}% applied to collections.")
        with dl_col:
            export_bytes = build_weekly_breakdown_export(weekly_results, selected_month)
            st.download_button(
                "Download Weekly Breakdown (Excel)", data=export_bytes,
                file_name=f"weekly_cashflow_{selected_month.strftime('%Y_%m')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
        with raw_col:
            monthly_raw_bytes = build_month_raw_export(weekly_results, selected_month)
            st.download_button(
                "Download Monthly Raw Data (Excel)", data=monthly_raw_bytes,
                file_name=f"monthly_raw_data_{selected_month.strftime('%Y_%m')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
        render_week_holiday_summary(weekly_results["week_meta"])
        render_weekly_financial_table(weekly_results, selected_month, balance_as_of_date)
        if resolve_weekly_drilldown_selection(weekly_results, selected_month) is None:
            render_weekly_drilldown_selection(weekly_results, selected_month)
        render_weekly_drilldown(weekly_results["week_meta"])

    # ══ SECTION 3: Collection Engine ═════════════════════════════════════════
    if selected_view == "Collection Engine":
        render_section_header(
            "Section 3: Collection Engine",
            f"Receivables view using configured probabilities (Base {b0_base:.0%} / Worst {b0_worst:.0%} for 0-30 days). TDS {tds_pct:.1f}% applied.",
        )
        render_collection_engine(scenario_table, overdue_dependency, sixty_plus_dependency)

    # ══ SECTION 4: 3-Month Cash Flow View ═══════════════════════════════════
    if selected_view == "3-Month Cash Flow View":
        render_section_header(
            "Section 4: 3-Month Cash Flow View",
            f"Monthly aggregation for forward liquidity monitoring. Cash floor: {format_currency(min_cash_threshold) if min_cash_threshold > 0 else 'Not set'}.",
        )
        render_six_month_view(
            weekly_results["monthly_forecast"],
            weekly_results["monthly_inflow_detail"],
            weekly_results["monthly_outflow_detail"],
            min_cash_threshold=min_cash_threshold,
        )

    # ══ SECTION 5: Actual vs Budget ══════════════════════════════════════════
    if selected_view == "Actual vs Budget":
        render_section_header(
            "Section 5: Actual vs Budget",
            "Load actuals in the same workbook structure to compare against plan and isolate cash-impacting variances.",
        )
        render_actual_vs_budget(
            weekly_results,
            actual_data,
            actual_file_name,
            base_probs=user_base_probs,
            worst_probs=user_worst_probs,
            tds_rate=tds_rate,
        )


    # ══ SECTION 6: Edit Data ═════════════════════════════════════════════════
    if selected_view == "✏️ Edit Data":
        _edit_window_start = weekly_results.get("selected_period_start", weekly_results["selected_month_start"])
        _edit_window_end = weekly_results["selected_month_end"]
        _edit_window_label = (
            f"Showing only rows relevant to the {month_label(selected_month)} cash flow window: "
            f"{format_date(_edit_window_start)} to {format_date(_edit_window_end)}."
        )
        render_section_header(
            "Edit Data — Live Override",
            "Edit receivable lines or outflow entries directly. Every change is reflected immediately "
            "in all other sections without re-uploading Excel. Track Changes keeps a separate undo/redo trail "
            "for inflows and outflows once edits are applied.",
        )

        _edit_badge = (
            "🟡 Working on edited data — edits active"
            if (
                not edit_frames_equal(st.session_state["edited_receivables"], _excel_receivables)
                or not edit_frames_equal(st.session_state["edited_outflows"], _excel_outflows)
            )
            else "🟢 Data matches the uploaded Excel"
        )
        st.caption(_edit_badge)
        try:
            reviewed_source_bytes = build_reuploadable_source_export(
                planned_upload,
                st.session_state["edited_receivables"],
                st.session_state["edited_outflows"],
                plan_data.get("workbook_layout"),
            )
        except Exception as exc:
            reviewed_source_bytes = None
            st.warning(f"Reviewed source export is unavailable right now: {exc}")

        if reviewed_source_bytes is not None:
            reviewed_file_name = f"{Path(planned_upload.name).stem}_reviewed_source.xlsx"
            st.download_button(
                "Download Reviewed Source (Excel)",
                data=reviewed_source_bytes,
                file_name=reviewed_file_name,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                help="Exports the current reviewed data back into upload-compatible sheet tabs so you can re-upload it later.",
            )

        tab_recv, tab_out = st.tabs(["📥 Inflows", "📤 Outflows"])

        # ── Tab 1: Receivables ────────────────────────────────────────────────
        with tab_recv:
            _recv_full_df = st.session_state["edited_receivables"].copy()
            _recv_forecast_view = derive_collection_probabilities(
                _recv_full_df,
                anchor_date=balance_as_of_date,
                base_probs=user_base_probs,
                worst_probs=user_worst_probs,
                tds_rate=tds_rate,
            )
            _recv_visible_mask = pd.to_datetime(
                _recv_forecast_view["forecast_collection_date"], errors="coerce"
            ).between(_edit_window_start, _edit_window_end, inclusive="both")
            _recv_visible_df = _recv_full_df.loc[_recv_visible_mask].copy()
            _recv_hidden_df = _recv_full_df.loc[~_recv_visible_mask].copy()
            st.markdown(
                f"**{len(_recv_visible_df)} receivable lines shown for {month_label(selected_month)}.** "
                "Edit amounts, dates, aging buckets or billing types. "
                "Add rows with ➕ or delete with the checkbox. "
                "Click **Apply** to commit changes."
            )
            st.caption(_edit_window_label)
            recv_display_cols = [
                "counterparty", "invoice_date", "tentative_collection_date",
                "due_date", "amount", "ageing", "aging_bucket", "billing_bucket",
            ]
            recv_edit_df = (
                _recv_visible_df
                .reindex(columns=recv_display_cols)
                .copy()
            )
            recv_edit_df.index = _recv_visible_df.index
            recv_edit_df = apply_sort_controls(
                recv_edit_df,
                key_prefix="edit_receivables",
                sortable_columns=recv_display_cols,
                preserve_index=True,
            )
            # Ensure date columns are proper Python date for the editor
            for _dc in ["invoice_date", "tentative_collection_date", "due_date"]:
                recv_edit_df[_dc] = pd.to_datetime(recv_edit_df[_dc], errors="coerce").dt.date

            edited_recv_df = st.data_editor(
                recv_edit_df,
                key="recv_data_editor",
                num_rows="dynamic",
                use_container_width=True,
                hide_index=True,
                column_config={
                    "counterparty": st.column_config.TextColumn(
                        "Client / Party", help="Customer or counterparty name"
                    ),
                    "invoice_date": st.column_config.DateColumn(
                        "Invoice Date", format="DD/MM/YYYY"
                    ),
                    "tentative_collection_date": st.column_config.DateColumn(
                        "Expected Collection Date",
                        format="DD/MM/YYYY",
                        help="Drive when this receivable lands in the weekly forecast",
                    ),
                    "due_date": st.column_config.DateColumn(
                        "Due Date", format="DD/MM/YYYY"
                    ),
                    "amount": st.column_config.NumberColumn(
                        "Amount (₹)", format="%.0f", min_value=0
                    ),
                    "ageing": st.column_config.NumberColumn(
                        "Ageing (days)", format="%.0f", min_value=0
                    ),
                    "aging_bucket": st.column_config.SelectboxColumn(
                        "Aging Bucket", options=["0-30", "30-60", "60+"]
                    ),
                    "billing_bucket": st.column_config.SelectboxColumn(
                        "Billing Type", options=["Advance", "Receivables"]
                    ),
                },
            )

            _rc1, _rc2, _rc3 = st.columns([2, 1, 1])
            with _rc1:
                if st.button("✅ Apply Receivables Changes", use_container_width=True, type="primary"):
                    # Rebuild full receivables df from the editor output
                    _new_recv = edited_recv_df.copy()
                    for _dc in ["invoice_date", "tentative_collection_date", "due_date"]:
                        _new_recv[_dc] = pd.to_datetime(_new_recv[_dc], errors="coerce").dt.normalize()
                    _new_recv["amount"] = pd.to_numeric(_new_recv["amount"], errors="coerce").fillna(0.0)
                    _new_recv["ageing"] = pd.to_numeric(_new_recv["ageing"], errors="coerce").fillna(0.0)
                    _new_recv["counterparty"] = _new_recv["counterparty"].fillna("Receivable").astype(str)
                    _new_recv["aging_bucket"] = _new_recv["aging_bucket"].fillna(
                        _new_recv["ageing"].apply(classify_aging_bucket)
                    )
                    _new_recv["billing_bucket"] = _new_recv["billing_bucket"].fillna("Receivables")
                    _new_recv["billing_bucket"] = _new_recv["billing_bucket"].map(normalize_billing_bucket)
                    # Rebuild description column to stay in sync
                    _new_recv["description"] = _new_recv["counterparty"]
                    # Drop zero-amount rows
                    _new_recv = _new_recv[_new_recv["amount"] != 0].copy()
                    # Restore non-editable columns for rows that remain visible after edits.
                    _hidden_cols = [
                        c for c in _recv_full_df.columns
                        if c not in recv_display_cols and c not in _new_recv.columns
                    ]
                    if _hidden_cols:
                        _existing_hidden = _recv_visible_df.reindex(columns=_hidden_cols)
                        _new_recv = _new_recv.join(_existing_hidden, how="left")
                    _orig_cols = [
                        c for c in _recv_full_df.columns
                        if c not in recv_display_cols
                    ]
                    for _oc in _orig_cols:
                        if _oc not in _new_recv.columns:
                            _new_recv[_oc] = None
                    if "source_sheet" in _new_recv.columns:
                        _default_inflow_sheet = plan_data.get("inflow_sheet_order", [None])[0] or "Receivables"
                        _new_recv["source_sheet"] = (
                            _new_recv["source_sheet"]
                            .fillna("")
                            .astype(str)
                            .str.strip()
                            .replace("", _default_inflow_sheet)
                        )
                    _new_recv = _new_recv.reindex(columns=_recv_full_df.columns, fill_value=None)
                    _final_recv = pd.concat([_recv_hidden_df, _new_recv], axis=0, sort=False)
                    _final_recv = _final_recv.sort_index(kind="stable").reset_index(drop=True)
                    _result = commit_edit_history("receivables", _final_recv, "Applied inflow changes")
                    if _result["changed"]:
                        st.success(
                            f"✅ Receivables updated — {len(_new_recv)} lines active. "
                            f"Tracked change: {_result['summary']}."
                        )
                        st.rerun()
                    st.info("No receivable changes detected to apply.")
            with _rc2:
                if st.button("🔄 Reset to Excel", use_container_width=True):
                    _result = commit_edit_history("receivables", _excel_receivables.copy(), "Reset inflows to Excel")
                    if _result["changed"]:
                        st.success("Receivables reset to original Excel data.")
                        st.rerun()
                    st.info("Receivables already match the uploaded Excel.")
            with _rc3:
                # Quick stats
                _total_recv = _recv_visible_df["amount"].sum() if not _recv_visible_df.empty else 0.0
                st.metric("Total Receivables (₹)", format_currency(_total_recv))

            render_track_changes_panel("receivables")

        # ── Tab 2: Outflows ───────────────────────────────────────────────────
        with tab_out:
            _out_full_df = st.session_state["edited_outflows"].copy()
            _out_visible_mask = pd.to_datetime(
                _out_full_df["date"], errors="coerce"
            ).between(_edit_window_start, _edit_window_end, inclusive="both")
            _out_visible_df = _out_full_df.loc[_out_visible_mask].copy()
            _out_hidden_df = _out_full_df.loc[~_out_visible_mask].copy()
            st.markdown(
                f"**{len(_out_visible_df)} outflow entries shown for {month_label(selected_month)}.** "
                "Edit amounts, dates, categories or add new rows. "
                "Click **Apply** to commit changes."
            )
            st.caption(_edit_window_label)
            out_display_cols = ["sheet_name", "date", "vendor_name", "description", "amount"]
            out_edit_df = (
                _out_visible_df
                .reindex(columns=out_display_cols)
                .copy()
            )
            out_edit_df.index = _out_visible_df.index
            out_edit_df = apply_sort_controls(
                out_edit_df,
                key_prefix="edit_outflows",
                sortable_columns=out_display_cols,
                preserve_index=True,
            )
            out_edit_df["date"] = pd.to_datetime(out_edit_df["date"], errors="coerce").dt.date

            # Build list of known categories for the selectbox
            _known_cats = sorted(set(
                plan_data["outflow_sheet_order"]
                + st.session_state["edited_outflows"]["sheet_name"].dropna().unique().tolist()
            ))

            edited_out_df = st.data_editor(
                out_edit_df,
                key="out_data_editor",
                num_rows="dynamic",
                use_container_width=True,
                hide_index=True,
                column_config={
                    "sheet_name": st.column_config.SelectboxColumn(
                        "Category",
                        options=_known_cats,
                        help="Outflow category / sheet (e.g. Employee Salaries, Vendor Payment)",
                    ),
                    "date": st.column_config.DateColumn(
                        "Payment Date", format="DD/MM/YYYY"
                    ),
                    "vendor_name": st.column_config.TextColumn(
                        "Vendor / Party", help="Payee name"
                    ),
                    "description": st.column_config.TextColumn(
                        "Description", help="Nature of payment or invoice reference"
                    ),
                    "amount": st.column_config.NumberColumn(
                        "Amount (₹)", format="%.0f", min_value=0
                    ),
                },
            )

            _oc1, _oc2, _oc3 = st.columns([2, 1, 1])
            with _oc1:
                if st.button("✅ Apply Outflow Changes", use_container_width=True, type="primary"):
                    _new_out = edited_out_df.copy()
                    _new_out["date"] = pd.to_datetime(_new_out["date"], errors="coerce").dt.normalize()
                    _new_out["amount"] = pd.to_numeric(_new_out["amount"], errors="coerce").fillna(0.0)
                    _new_out["vendor_name"] = _new_out["vendor_name"].fillna("").astype(str)
                    _new_out["description"] = _new_out["description"].fillna("").astype(str)
                    _new_out["sheet_name"] = _new_out["sheet_name"].fillna("Other")
                    # Drop rows without date or amount
                    _new_out = _new_out[
                        _new_out["date"].notna() & (_new_out["amount"] != 0)
                    ].copy()
                    _new_out = _new_out.reindex(columns=_out_full_df.columns, fill_value=None)
                    _final_out = pd.concat([_out_hidden_df, _new_out], axis=0, sort=False)
                    _final_out = _final_out.sort_index(kind="stable").reset_index(drop=True)
                    _result = commit_edit_history("outflows", _final_out, "Applied outflow changes")
                    if _result["changed"]:
                        st.success(
                            f"✅ Outflows updated — {len(_new_out)} entries active. "
                            f"Tracked change: {_result['summary']}."
                        )
                        st.rerun()
                    st.info("No outflow changes detected to apply.")
            with _oc2:
                if st.button("🔄 Reset to Excel ", use_container_width=True):
                    _result = commit_edit_history("outflows", _excel_outflows.copy(), "Reset outflows to Excel")
                    if _result["changed"]:
                        st.success("Outflows reset to original Excel data.")
                        st.rerun()
                    st.info("Outflows already match the uploaded Excel.")
            with _oc3:
                _total_out = _out_visible_df["amount"].sum() if not _out_visible_df.empty else 0.0
                st.metric("Total Outflows (₹)", format_currency(_total_out))

            render_track_changes_panel("outflows")

        st.markdown("---")
        st.markdown("#### Live Data Summary (post-edit)")
        _plan_summary = weekly_results["selected_month_plan"]
        _ending_excl_od = _plan_summary.get("ending_excl_od", (_plan_summary.get("ending", 0.0) or 0.0))
        _od_utilization = _plan_summary.get("od_utilization", 0.0) or 0.0

        _sum_c1, _sum_c2, _sum_c3, _sum_c4 = st.columns(4)
        with _sum_c1:
            st.metric(
                "Receivable lines",
                len(st.session_state["edited_receivables"]),
                delta=len(st.session_state["edited_receivables"]) - len(_excel_receivables) or None,
            )
        with _sum_c2:
            st.metric(
                "Outflow entries",
                len(st.session_state["edited_outflows"]),
                delta=len(st.session_state["edited_outflows"]) - len(_excel_outflows) or None,
            )
        with _sum_c3:
            st.metric(
                "Ending Balance Excl. OD (₹)",
                format_currency(_ending_excl_od),
            )
        with _sum_c4:
            st.metric(
                "OD Utilization (₹)",
                format_currency(_od_utilization),
                delta=None if _od_utilization > 0 else "No OD",
            )


if __name__ == "__main__":
    main()
